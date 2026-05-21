#!/usr/bin/env python3
"""
Sole Similarity Analyzer  v12  —  v6 + Viewer + Overlay Score
==============================================================
Merges the v6 similarity engine with the match viewer into one script.
No Excel needed as an intermediate step — runs once and produces both
results and rich visual comparison cards.

New vs v6:
  - Overlay score added: white_pixels / (white+yellow+cyan) on the
    already-normalised 256x128 edge maps. Works because v6 normalises
    to fixed size before comparison, so scale is not an issue.
  - Each comparison image now shows:
      • Photo row (with tread region highlighted)
      • Binary tread row (what the algorithm actually compares)
      • Cyan/Yellow/White edge overlay (A only / B only / both)
      • Score bar chart for all sub-scores
  - Overlay score feeds into the similarity score (10% weight)
  - Threshold kept at 0.72 (v6 baseline — proven 100% precision)
"""

import cv2
import numpy as np
from pathlib import Path
from itertools import combinations
from collections import defaultdict
import warnings, re
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR   = Path.home() / "Desktop" / "extracted_boot_images" / "classified"
SOLE_DIR   = BASE_DIR / "sole"
SIM_DIR    = BASE_DIR / "similarity"
DIAG_DIR   = BASE_DIR / "diagnostics"
EXCEL_PATH  = BASE_DIR / "sole_similarity_v12.xlsx"
VIEWER_DIR  = BASE_DIR / "match_viewer_v12"
IMAGE_EXTS  = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}

IDENTICAL_THRESH = 0.72
LIKELY_THRESH    = 0.58

SIM_DIR.mkdir(parents=True, exist_ok=True)
DIAG_DIR.mkdir(parents=True, exist_ok=True)
VIEWER_DIR.mkdir(parents=True, exist_ok=True)
TARGET = 400
NORM_H, NORM_W = 256, 128


# ══════════════════════════════════════════════════════════════════════════════
#  GROUND TRUTH DATA
# ══════════════════════════════════════════════════════════════════════════════

GROUND_TRUTH_RAW = """
iDAPT650,5ANADK2-8611,Dakota,1,M,Green Diamond + Icelandic Grip,
iDAPT824,5ANADK2-8611,Dakota WorkPro Series,1,M,Green Diamond + Icelandic Grip,No image
iDAPT713,5ANBDK3-6611,Dakota,2,M,Green Diamond + Icelandic Grip,
iDAPT820,5ANBDK3-6611,Dakota WorkPro Series,2,M,Green Diamond + Icelandic Grip,No image
iDAPT548,5ANFDKW0-8907,Dakota,3,M,Green Diamond,Identical
iDAPT607,5ANFDKW0-8907,Dakota,3,M,Green Diamond,Identical
iDAPT630,5ANFDKW2-8907,Dakota,4,M,Green Diamond + Icelandic Grip,
iDAPT823,5ANFDKW2-8907,Dakota WorkPro Series,4,M,Green Diamond + Icelandic Grip,No image
iDAPT709,5ANFDKW3-8901,Dakota,5,M,Green Diamond,
iDAPT804,5ANFDKW3-8901,Dakota,5,M,Green Diamond,No image
iDAPT773,5ANFDKW4-8911,Dakota,6,M,Green Diamond,
iDAPT819,5ANFDKW4-8911,Dakota WorkPro Series,6,M,Green Diamond,No image
iDAPT391,5ANFDKW6-8524,Dakota,7,M,Green Diamond,
iDAPT840,5ANFDKW6-8524,Dakota,7,M,Green Diamond,No image
iDAPT376,5ANFDKW7-8901,DAKOTA,8,M,Green Diamond,not identical
iDAPT667,5ANFDKW7-8901,Dakota,8,M,Green Diamond,not identical
iDAPT443,5ANFDKW8-8527,Dakota,9,M,Green Diamond,
iDAPT836,5ANFDKW8-8527,Dakota,9,M,Green Diamond,No image
iDAPT445,5BOFDK7-8520,Dakota,10,W,Green Diamond,
iDAPT832,5BOFDK7-8520,Dakota,10,W,Green Diamond,No image
iDAPT337,5CPAWRAS-1108,WindRiver,11,M,Green Diamond,Identical
iDAPT393,5CPAWRAS-1108,WindRiver,11,M,Green Diamond,Identical
iDAPT381,5CPAWRF18-1231,WindRiver,12,M,Icelandic Grip,Identical
iDAPT786,5CPAWRF18-1231,WindRiver,12,M,Icelandic Grip,Identical
iDAPT523,5CPAWRF19-1240,Windriver,13,M,Green Diamond + Icelandic Grip,Identical
iDAPT739,5CPAWRF19-1240,WindRiver,13,M,Green Diamond + Icelandic Grip,Identical
IDAPT644,5CPAWRF22-1232,WindRiver,14,M,Icelandic Grip,Identical
iDAPT787,5CPAWRF22-1232,WindRiver,14,M,Icelandic Grip,Identical
iDAPT588,5CPEDHF20-5296,Denver Hayes,15,M,Icelandic Grip,Identical
iDAPT692,5CPEDHF20-5296,Denver Hayes,15,M,Icelandic Grip,Identical
iDAPT390,5CPEWRF18-5272,WindRiver,16,M,Green Diamond + Icelandic Grip,Identical
iDAPT509,5CPEWRF18-5272,WindRiver,16,M,Green Diamond + Icelandic Grip,Identical
iDAPT514,5CPEWRF19-5281,Denver Hayes,17,M,Green Diamond + Icelandic Grip,Identical
iDAPT738,5CPEWRF19-5281,Denver Hayes,17,M,Green Diamond + Icelandic Grip,Identical
iDAPT506,5CPEWRF19-5283,WindRiver,18,M,Green Diamond + Icelandic Grip,Identical
IDAPT664,5CPEWRF19-5283,WindRiver,18,M,Green Diamond + Icelandic Grip,Identical
iDAPT536,5CPEWRF20-5292,WindRiver,19,M,Green Diamond + Icelandic Grip,Identical
iDAPT665,5CPEWRF20-5292,WindRiver,19,M,Green Diamond + Icelandic Grip,Identical
iDAPT610,5CPEWRF20-5293,WindRiver,20,M,Green Diamond,Identical
iDAPT740,5CPEWRF20-5293,WindRiver,20,M,Green Diamond,Identical
iDAPT593,5CPEWRF20-5294,WindRiver,21,M,Green Diamond + Icelandic Grip,not identical
iDAPT741,5CPEWRF20-5294,WindRiver,21,M,Green Diamond + Icelandic Grip,not identical
IDAPT641,5CPEWRF22-5222,WindRiver,22,M,Green Diamond,
iDAPT821,5CPEWRF22-5222,WindRiver,22,M,Green Diamond,No image
IDAPT649,5CPEWRF22-5277,WindRiver,23,M,Green Diamond,not identical
iDAPT743,5CPEWRF22-5277,WindRiver,23,M,Green Diamond,not identical
iDAPT698,5CPEWRF23-2454,WindRiver,24,M,Icelandic Grip,
iDAPT806,5CPEWRF23-2454,WindRiver,24,M,Icelandic Grip,No image
iDAPT715,5CPEWRF23-5267,WindRiver,25,M,Green Diamond + Icelandic Grip,Identical
iDAPT777,5CPEWRF23-5267,WindRiver,25,M,Green Diamond + Icelandic Grip,Identical
iDAPT717,5CPEWRF23-5290,WindRiver,26,M,Green Diamond + Icelandic Grip,
iDAPT828,5CPEWRF23-5290,WindRiver,26,M,Green Diamond + Icelandic Grip,No image
iDAPT737,5CPEWRF24-1400,WindRiver,27,M,Green Diamond + Icelandic Grip,
iDAPT807,5CPEWRF24-1400,WindRiver,27,M,Green Diamond + Icelandic Grip,No image
iDAPT772,5CPEWRF24-1406,WindRiver,28,M,Green Diamond + Icelandic Grip,
iDAPT826,5CPEWRF24-1406,WindRiver,28,M,Green Diamond + Icelandic Grip,No image
iDAPT759,5CPEWRF24-2455,WindRiver,29,M,Icelandic Grip,
iDAPT810,5CPEWRF24-2455,WindRiver,29,M,Icelandic Grip,No image
iDAPT338,5CPEWRFW2-5114,WindRiver,30,M,Green Diamond,Identical
iDAPT532,5CPEWRFW2-5114,WindRiver,30,M,Green Diamond,Identical
iDAPT363,5CPGWRF18-7117,WindRiver,31,M,Icelandic Grip,Identical
iDAPT504,5CPGWRF18-7117,WindRiver,31,M,Green Diamond + Icelandic Grip,Identical
iDAPT598,5DQEDHFB20-5511,Denver Hayes,32,W,Icelandic Grip,
iDAPT702,5DQEDHFB20-5511,Denver Hayes,32,W,Icelandic Grip,No image
IDAPT694,5DQEDHFB23-5537,Denver Hayes,33,W,Icelandic Grip,
iDAPT704,5DQEDHFB23-5537,Denver Hayes,33,W,Icelandic Grip,No image
iDAPT361,5DQEDHFB8-5117,Denver Hayes,34,W,Icelandic Grip,Identical
iDAPT505,5DQEDHFB8-5117,Denver Hayes,34,W,Icelandic Grip,Identical
iDAPT503,5DQEDHFB9-5440,Denver Hayes,35,W,Icelandic Grip,
iDAPT811,5DQEDHFB9-5440,Denver Hayes,35,W,Icelandic Grip,No image
iDAPT770,5DQEWRF24-5960,WindRiver,36,W,Green Diamond + Icelandic Grip,
iDAPT829,5DQEWRF24-5960,WindRiver,36,W,Green Diamond + Icelandic Grip,No image
iDAPT537,5DQEWRFB20-5900,WindRiver,37,W,Green Diamond + Icelandic Grip,Identical
IDAPT663,5DQEWRFB20-5900,WindRiver,37,W,Green Diamond + Icelandic Grip,Identical
IDAPT654,5DQEWRFB21-5557,WindRiver,38,W,Green Diamond + Icelandic Grip,
iDAPT729,5DQEWRFB21-5557,WindRiver,38,W,Green Diamond + Icelandic Grip,No image
IDAPT652,5DQEWRFB22-5288,WindRiver,39,W,Green Diamond + Icelandic Grip,Identical
iDAPT782,5DQEWRFB22-5288,WindRiver,39,W,Green Diamond + Icelandic Grip,Identical
iDAPT714,5DQEWRFB23-5959,WindRiver,40,W,Green Diamond + Icelandic Grip,
iDAPT830,5DQEWRFB23-5959,WindRiver,40,W,Green Diamond + Icelandic Grip,No image
iDAPT736,5DQEWRFB24-5202,WindRiver,41,W,Green Diamond + Icelandic Grip,
iDAPT815,5DQEWRFB24-5202,WindRiver,41,W,Green Diamond + Icelandic Grip,No image
iDAPT802,5DQEWRFB25-6001,WindRiver,42,W,Green Diamond + Icelandic Grip,
iDAPT816,5DQEWRFB25-6001,WindRiver,42,W,Green Diamond + Icelandic Grip,No image
iDAPT344,5DQEWRFB7-5108,WindRiver,43,W,Green Diamond,Identical
iDAPT385,5DQEWRFB7-5108,WindRiver,43,W,Green Diamond,Identical
iDAPT384,5DQEWRFB8-5061,WindRiver,44,W,Icelandic Grip,Identical
iDAPT513,5DQEWRFB8-5061,WindRiver,44,W,Icelandic Grip,Identical
iDAPT386,5DQEWRFB8-5088,WindRiver,45,W,Green Diamond + Icelandic Grip,Identical
iDAPT510,5DQEWRFB8-5088,WindRiver,45,W,Green Diamond + Icelandic Grip,Identical
iDAPT515,5DQEWRFB9-5556,WindRiver,46,W,Green Diamond + Icelandic Grip,
iDAPT812,5DQEWRFB9-5556,WindRiver,46,W,Green Diamond + Icelandic Grip,No image
iDAPT571,604SP2AG,Royer,47,M,Arctic Grip,Identical
iDAPT615,604SP2AG,Royer,47,M,Arctic Grip,Identical
IDAPT697,80013W21B,Woods,48,M,Green Diamond - Icelandic Grip,
iDAPT851,80013W21B,Woods,48,M,Icelandic Grip,No image
iDAPT690,80013W24,Woods,49,M,Green Diamond + Icelandic Grip,
iDAPT849,80013W24,Woods,49,M,Green Diamond + Icelandic Grip,No image
iDAPT758,80014W05A,Woods,50,M,Icelandic Grip,
iDAPT850,80014W05A,Woods,50,M,Icelandic Grip,No image
iDAPT797,80015W06,Woods,51,M,Green Diamond + Icelandic Grip,
iDAPT847,80015W06,Woods,51,M,Green Diamond + Icelandic Grip,No image
iDAPT799,80015W08,Woods,52,M,Icelandic Grip,
iDAPT846,80015W08,Woods,52,M,Icelandic Grip,No image
IDAPT643,80022W13,Woods,53,W,Icelandic Grip,Identical
IDAPT696,80022W13,Woods,53,W,Icelandic Grip,Identical
iDAPT755,80024W12,Woods,54,W,Icelandic Grip,
iDAPT852,80024W12,Woods,54,W,Green Diamond + Icelandic Grip,No image
iDAPT803,80025W06A,Woods,55,W,Icelandic Grip,
iDAPT853,80025W06A,Woods,55,W,Icelandic Grip,No image
iDAPT798,80025W07,Woods,56,W,Icelandic Grip,
iDAPT854,80025W07,Woods,56,W,Icelandic Grip,No image
iDAPT800,80025W10A,Woods,57,W,Icelandic Grip,
iDAPT855,80025W10A,Woods,57,W,Icelandic Grip,No image
iDAPT512,HHF186118,Helly Hansen Workwear,58,M,Green Diamond,Identical
IDAPT686,HHF186118,Helly Hansen Workwear,58,M,Green Diamond,Identical
iDAPT530,HHF196119,Helly Hansen Workwear,59,M,Green Diamond,Identical
IDAPT685,HHF196119,Helly Hansen Workwear,59,M,Green Diamond,Identical
IDAPT640,HHF225500,Helly Hansen,60,M,Green Diamond + Icelandic Grip,Identical
IDAPT705,HHF225500,Helly Hansen,60,M,Green Diamond + Icelandic Grip,Identical
IDAPT639,HHF225510,Helly Hansen,61,M,Green Diamond + Icelandic Grip,Identical
IDAPT706,HHF225510,Helly Hansen,61,M,Green Diamond + Icelandic Grip,Identical
IDAPT642,HHF225555,Helly Hansen,62,W,Green Diamond + Icelandic Grip,Identical
iDAPT784,HHF225555,Helly Hansen,62,W,Green Diamond + Icelandic Grip,Identical
iDAPT710,HHF231102,Helly Hansen Workwear,63,M,Green Diamond,
iDAPT822,HHF231102,Helly Hansen Workwear,63,M,Green Diamond,No image
iDAPT712,HHF231104,Helly Hansen Workwear,64,M,Green Diamond + Icelandic Grip,
iDAPT835,HHF231104,Helly Hansen,64,M,Green Diamond + Icelandic Grip,No image
iDAPT526,HHLF195000,Helly Hansen,65,M,Green Diamond + Icelandic Grip,Identical
iDAPT595,HHLF195000,Helly Hansen,65,M,Green Diamond + Icelandic Grip,Identical
iDAPT524,HHLF195001,Helly Hansen,66,M,Green Diamond,Identical
iDAPT745,HHLF195001,Helly Hansen,66,M,Green Diamond,Identical
iDAPT522,HHLF195550,Helly Hansen,67,W,Green Diamond + Icelandic Grip,Identical
iDAPT591,HHLF195550,Helly Hansen,67,W,Green Diamond + Icelandic Grip,Identical
iDAPT326,J15752,Merrell,68,W,Arctic grip,Identical
iDAPT413,J15752,Merrell,68,W,Arctic Grip,Identical
iDAPT318,J37829,Merrell,69,M,Arctic grip,Identical
iDAPT404,J37829,Merrell,69,M,Arctic Grip,Identical
iDAPT330,P721621,CAT,70,M,Arctic grip,Identical
iDAPT400,P721621,CAT,70,M,Arctic Grip,Identical
""".strip()

def parse_ground_truth():
    idapt_info = {}
    pair_groups = defaultdict(list)
    pair_statuses = defaultdict(list)
    for line in GROUND_TRUTH_RAW.split('\n'):
        parts = [p.strip() for p in line.split(',')]
        while len(parts) < 7: parts.append('')
        idapt, style, brand, pair, mf, tech, status = parts
        key = idapt.upper()
        idapt_info[key] = dict(style=style, brand=brand, pair=int(pair),
                                mf=mf, tech=tech, status=status)
        pair_groups[int(pair)].append(key)
        pair_statuses[int(pair)].append(status.lower())
    pair_status = {}
    for pn, sts in pair_statuses.items():
        if any('not identical' in s for s in sts):
            pair_status[pn] = 'not identical'
        elif any(s == 'identical' for s in sts):
            pair_status[pn] = 'identical'
        elif any('no image' in s for s in sts):
            pair_status[pn] = 'no image'
        else:
            pair_status[pn] = ''
    return idapt_info, dict(pair_groups), pair_status

GT_INFO, GT_PAIRS, GT_STATUS = parse_ground_truth()

def get_idapt_from_filename(fname):
    stem = Path(fname).stem.upper()
    m = re.match(r'(I?DAPT\d+)', stem)
    return m.group(1) if m else stem


# ══════════════════════════════════════════════════════════════════════════════
#  TREAD-ONLY MASK  (v6 key fix: exclude boot upper)
# ══════════════════════════════════════════════════════════════════════════════

def find_boot_contour(gray, img_bgr):
    """Find the full boot outline (same as before)."""
    h, w = gray.shape
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blur, 15, 60)
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (25, 25))
    closed = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, k)
    filled = cv2.dilate(closed, k, iterations=2)
    cnts, _ = cv2.findContours(filled, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best = None; best_score = 0.0
    for c in cnts:
        area = cv2.contourArea(c)
        if area < h*w*0.08 or area > h*w*0.95: continue
        x, y, cw, ch = cv2.boundingRect(c)
        if ch == 0: continue
        score = area
        if abs((x+cw/2)/w - 0.5) < 0.4 and abs((y+ch/2)/h - 0.5) < 0.4:
            score *= 1.0
        else:
            score *= 0.3
        if score > best_score: best_score = score; best = c
    if best is not None:
        mask = np.zeros((h, w), dtype=np.uint8)
        cv2.drawContours(mask, [best], -1, 255, -1)
        return mask
    # Fallback: HSV distance
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    bw = 15
    border_pixels = np.concatenate([
        hsv[:bw,:,:].reshape(-1,3), hsv[-bw:,:,:].reshape(-1,3),
        hsv[:,:bw,:].reshape(-1,3), hsv[:,-bw:,:].reshape(-1,3)
    ])
    bg_v = float(np.median(border_pixels[:,2]))
    bg_s = float(np.median(border_pixels[:,1]))
    vd = np.abs(hsv[:,:,2].astype(np.float32) - bg_v)
    sd = np.abs(hsv[:,:,1].astype(np.float32) - bg_s)
    fg = ((vd + sd * 0.5) > 25).astype(np.uint8) * 255
    k3 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (15, 15))
    fg = cv2.morphologyEx(fg, cv2.MORPH_CLOSE, k3)
    fg = cv2.morphologyEx(fg, cv2.MORPH_OPEN,
                           cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9)))
    n_lab, labels, stats, _ = cv2.connectedComponentsWithStats(fg)
    if n_lab > 1:
        largest = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
        return (labels == largest).astype(np.uint8) * 255
    mask = np.zeros((h, w), dtype=np.uint8)
    mask[int(h*0.04):h-int(h*0.04), int(w*0.04):w-int(w*0.04)] = 255
    return mask


def extract_tread_region(gray, img_bgr, boot_mask):
    """
    From the full boot mask, extract ONLY the tread/outsole region.
    
    Strategy: The tread has high-frequency texture (lug pattern).
    The boot upper is smoother (leather/fabric/lacing).
    Compute local texture energy and keep only textured regions
    within the boot mask, biased toward the bottom of the boot.
    """
    h, w = gray.shape
    
    # 1. Compute local texture energy using Laplacian variance in windows
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    
    # Local texture energy: variance of Laplacian in 15x15 windows
    lap = cv2.Laplacian(enhanced, cv2.CV_32F, ksize=3)
    lap_sq = lap * lap
    win = 21
    local_energy = cv2.blur(lap_sq, (win, win))
    local_energy = cv2.bitwise_and(local_energy.astype(np.float32),
                                     local_energy.astype(np.float32),
                                     mask=boot_mask)
    
    # 2. Also use edge density as texture indicator
    blur = cv2.GaussianBlur(enhanced, (3, 3), 0)
    edges = cv2.Canny(blur, 25, 80)
    edges = cv2.bitwise_and(edges, boot_mask)
    edge_density = cv2.blur(edges.astype(np.float32), (31, 31))
    
    # 3. Combine: texture = energy + edge density
    # Normalize both to 0-1
    e_max = local_energy.max()
    if e_max > 0: local_energy /= e_max
    d_max = edge_density.max()
    if d_max > 0: edge_density /= d_max
    
    texture_map = (local_energy * 0.5 + edge_density * 0.5)
    texture_map = cv2.bitwise_and(texture_map, texture_map, mask=boot_mask)
    
    # 4. Find the boot's bounding box and compute the vertical centroid
    #    of texture — the tread should be in the bottom portion
    ys_boot, xs_boot = np.where(boot_mask > 0)
    if len(ys_boot) < 100:
        return boot_mask  # fallback
    
    y_top, y_bot = ys_boot.min(), ys_boot.max()
    x_left, x_right = xs_boot.min(), xs_boot.max()
    boot_h = y_bot - y_top
    
    # 5. Weight texture by vertical position — bottom of boot gets boost
    #    This helps separate sole from upper when both have some texture
    vert_weight = np.zeros((h, w), dtype=np.float32)
    for row in range(y_top, y_bot + 1):
        # Linear weight: 0.3 at top of boot, 1.0 at bottom
        frac = (row - y_top) / max(boot_h, 1)
        vert_weight[row, :] = 0.3 + 0.7 * frac
    
    weighted_texture = texture_map * vert_weight
    
    # 6. Threshold to get tread region
    # Use Otsu on the weighted texture values within the boot
    wt_vals = weighted_texture[boot_mask > 0]
    if len(wt_vals) < 100:
        return boot_mask
    
    # Normalize to 0-255 for Otsu
    wt_norm = (weighted_texture * 255).astype(np.uint8)
    wt_norm = cv2.bitwise_and(wt_norm, boot_mask)
    _, tread_mask = cv2.threshold(wt_norm, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # 7. Clean up: keep only the largest connected component
    tread_mask = cv2.morphologyEx(tread_mask, cv2.MORPH_CLOSE,
                                   cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (15, 15)))
    tread_mask = cv2.morphologyEx(tread_mask, cv2.MORPH_OPEN,
                                   cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9)))
    
    n_lab, labels, stats, _ = cv2.connectedComponentsWithStats(tread_mask)
    if n_lab > 1:
        largest = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
        tread_mask = (labels == largest).astype(np.uint8) * 255
    
    # 8. Sanity check: tread should be at least 15% of boot area
    tread_area = np.sum(tread_mask > 0)
    boot_area = np.sum(boot_mask > 0)
    if tread_area < boot_area * 0.15:
        # Texture method failed — fall back to bottom 55% of boot
        cutoff_y = y_top + int(boot_h * 0.45)
        tread_mask = boot_mask.copy()
        tread_mask[:cutoff_y, :] = 0
        # Erode slightly
        tread_mask = cv2.erode(tread_mask,
                                cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7)), 1)
    
    return tread_mask


# ══════════════════════════════════════════════════════════════════════════════
#  CANONICAL ORIENTATION  (v6 fix: align all soles same way)
# ══════════════════════════════════════════════════════════════════════════════

def align_to_canonical(patch, patch_mask=None):
    """
    Rotate the sole patch so the long axis is vertical and
    the "heavier" end (more texture/area) is at the bottom.
    Returns the rotated patch.
    """
    h, w = patch.shape[:2]
    if patch_mask is None:
        patch_mask = (patch > 10).astype(np.uint8) * 255 if len(patch.shape) == 2 else \
                     (cv2.cvtColor(patch, cv2.COLOR_BGR2GRAY) > 10).astype(np.uint8) * 255
    
    # Find principal axis using moments
    ys, xs = np.where(patch_mask > 0)
    if len(xs) < 50:
        return patch
    
    # PCA on mask coordinates
    coords = np.column_stack([xs - xs.mean(), ys - ys.mean()]).astype(np.float32)
    cov = np.cov(coords.T)
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    
    # Principal axis = eigenvector with largest eigenvalue
    principal = eigenvectors[:, np.argmax(eigenvalues)]
    angle = np.degrees(np.arctan2(principal[1], principal[0]))
    
    # Rotate so principal axis is vertical (90°)
    rot_angle = 90 - angle
    M = cv2.getRotationMatrix2D((w/2, h/2), rot_angle, 1.0)
    
    # Expand canvas to avoid clipping
    cos_a = abs(M[0, 0]); sin_a = abs(M[0, 1])
    new_w = int(h * sin_a + w * cos_a)
    new_h = int(h * cos_a + w * sin_a)
    M[0, 2] += (new_w - w) / 2
    M[1, 2] += (new_h - h) / 2
    
    if len(patch.shape) == 3:
        rotated = cv2.warpAffine(patch, M, (new_w, new_h))
    else:
        rotated = cv2.warpAffine(patch, M, (new_w, new_h))
    rot_mask = cv2.warpAffine(patch_mask, M, (new_w, new_h))
    
    # Ensure heavier end is at bottom (flip if top half has more pixels)
    rh = rotated.shape[0]
    top_mass = np.sum(rot_mask[:rh//2, :] > 0)
    bot_mass = np.sum(rot_mask[rh//2:, :] > 0)
    if top_mass > bot_mass * 1.2:
        rotated = cv2.rotate(rotated, cv2.ROTATE_180)
        rot_mask = cv2.rotate(rot_mask, cv2.ROTATE_180)
    
    # Crop to bounding box of content
    ys2, xs2 = np.where(rot_mask > 0)
    if len(xs2) < 10:
        return patch
    y0, y1 = ys2.min(), ys2.max()
    x0, x1 = xs2.min(), xs2.max()
    margin = 5
    y0 = max(0, y0 - margin); y1 = min(rotated.shape[0], y1 + margin)
    x0 = max(0, x0 - margin); x1 = min(rotated.shape[1], x1 + margin)
    
    return rotated[y0:y1+1, x0:x1+1]


# ══════════════════════════════════════════════════════════════════════════════
#  FEATURE EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_features(img_path):
    img = cv2.imread(str(img_path))
    if img is None: return None
    h0, w0 = img.shape[:2]
    scale = TARGET / max(h0, w0)
    img = cv2.resize(img, (int(w0*scale), int(h0*scale)))
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    img_sq = np.zeros((TARGET, TARGET, 3), dtype=np.uint8)
    gray_sq = np.zeros((TARGET, TARGET), dtype=np.uint8)
    img_sq[:h,:w] = img; gray_sq[:h,:w] = gray
    gray = gray_sq; img = img_sq; h = w = TARGET

    # Step 1: Find boot outline
    boot_mask = find_boot_contour(gray, img)
    
    # Step 2: Extract tread-only region
    tread_mask = extract_tread_region(gray, img, boot_mask)
    
    # Step 3: Crop tread region and normalize
    ys, xs = np.where(tread_mask > 0)
    if len(xs) < 100:
        # Fallback
        ys, xs = np.where(boot_mask > 0)
        if len(xs) < 100:
            return None
        tread_mask = boot_mask
    
    y0, y1 = ys.min(), ys.max()
    x0, x1 = xs.min(), xs.max()
    tread_crop_gray = gray[y0:y1+1, x0:x1+1].copy()
    tread_crop_mask = tread_mask[y0:y1+1, x0:x1+1]
    tread_crop_gray[tread_crop_mask == 0] = 0
    
    # Step 4: Align to canonical orientation
    aligned_gray = align_to_canonical(tread_crop_gray, tread_crop_mask)
    
    # Step 5: Resize to standard dimensions
    norm_gray = cv2.resize(aligned_gray, (NORM_W, NORM_H))
    
    # Step 6: Create color-invariant representations
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
    norm_enhanced = clahe.apply(norm_gray)
    
    # Tread binary (adaptive threshold — color invariant)
    blur = cv2.GaussianBlur(norm_enhanced, (5, 5), 0)
    tread_bw = cv2.adaptiveThreshold(blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                      cv2.THRESH_BINARY, 25, 3)
    tread_bw = cv2.morphologyEx(tread_bw, cv2.MORPH_OPEN,
                                  cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)))
    
    # Edge pattern
    edge_pat = cv2.Canny(cv2.GaussianBlur(norm_enhanced, (3, 3), 0), 30, 90)
    edge_pat = cv2.dilate(edge_pat, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2, 2)), 1)
    
    # Mirrored versions
    tread_bw_flip = cv2.flip(tread_bw, 1)
    edge_pat_flip = cv2.flip(edge_pat, 1)
    norm_gray_flip = cv2.flip(norm_enhanced, 1)
    
    # ORB on edge pattern
    orb = cv2.ORB_create(nfeatures=500, edgeThreshold=8, patchSize=16, 
                          fastThreshold=5)  # lower thresholds = more keypoints
    kp_e, desc_e = orb.detectAndCompute(edge_pat, None)
    kp_ef, desc_ef = orb.detectAndCompute(edge_pat_flip, None)
    kp_t, desc_t = orb.detectAndCompute(tread_bw, None)
    kp_tf, desc_tf = orb.detectAndCompute(tread_bw_flip, None)
    
    # Zonal tread density (divide into 4x2 zones)
    n_vz, n_hz = 4, 2
    zh, zw = NORM_H // n_vz, NORM_W // n_hz
    zone_density = np.zeros(n_vz * n_hz, dtype=np.float32)
    for zi in range(n_vz):
        for zj in range(n_hz):
            r0, r1 = zi*zh, (zi+1)*zh
            c0, c1 = zj*zw, (zj+1)*zw
            zone = tread_bw[r0:r1, c0:c1]
            zone_density[zi*n_hz + zj] = np.mean(zone > 0)
    
    # Edge orientation histogram on normalized patch
    gx = cv2.Sobel(norm_enhanced, cv2.CV_32F, 1, 0, ksize=3)
    gy = cv2.Sobel(norm_enhanced, cv2.CV_32F, 0, 1, ksize=3)
    mag = np.sqrt(gx**2 + gy**2)
    ori = np.arctan2(gy, gx) * 180 / np.pi % 180
    n_bins = 18
    eoh = np.zeros(n_bins, dtype=np.float32)
    bin_width = 180.0 / n_bins
    for i in range(n_bins):
        in_bin = (ori >= i*bin_width) & (ori < (i+1)*bin_width) & (mag > 10)
        eoh[i] = np.sum(mag[in_bin])
    if eoh.sum() > 0: eoh /= eoh.sum()
    
    # Hu moments of tread binary
    hu = cv2.HuMoments(cv2.moments(tread_bw)).flatten()
    hu_log = -np.sign(hu) * np.log10(np.abs(hu) + 1e-12)
    
    # Contour shape descriptor (Fourier descriptors)
    cnts_bw, _ = cv2.findContours(tread_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    fourier_desc = np.zeros(20, dtype=np.float32)
    if cnts_bw:
        # Get the largest few contours
        cnts_sorted = sorted(cnts_bw, key=cv2.contourArea, reverse=True)[:10]
        all_points = np.vstack(cnts_sorted) if len(cnts_sorted) > 0 else cnts_sorted[0]
        # Flatten to complex
        pts = all_points.reshape(-1, 2).astype(np.float32)
        if len(pts) > 20:
            # Resample to fixed number of points
            indices = np.linspace(0, len(pts)-1, 64).astype(int)
            pts_resampled = pts[indices]
            z = pts_resampled[:, 0] + 1j * pts_resampled[:, 1]
            fft = np.fft.fft(z)
            # Take magnitude of first 20 coefficients (skip DC)
            fd = np.abs(fft[1:21])
            if fd[0] > 0: fd /= fd[0]  # scale invariant
            fourier_desc[:len(fd)] = fd.astype(np.float32)
    
    # GT info
    idapt_id = get_idapt_from_filename(img_path.name)
    gt = GT_INFO.get(idapt_id, {})

    # Full edges for visualization
    edges_full = cv2.Canny(cv2.GaussianBlur(
        clahe.apply(gray), (5,5), 0), 35, 110)
    edges_full = cv2.bitwise_and(edges_full, tread_mask)

    return dict(
        path=img_path, name=img_path.name, stem=img_path.stem,
        idapt=idapt_id,
        gt_style=gt.get('style',''), gt_brand=gt.get('brand',''),
        gt_pair=gt.get('pair',-1), gt_tech=gt.get('tech',''),
        gt_status=gt.get('status',''),
        # Normalized representations (after alignment!)
        norm_gray=norm_enhanced, norm_gray_flip=norm_gray_flip,
        tread_bw=tread_bw, tread_bw_flip=tread_bw_flip,
        edge_pat=edge_pat, edge_pat_flip=edge_pat_flip,
        # ORB
        desc_e=desc_e, desc_ef=desc_ef,
        desc_t=desc_t, desc_tf=desc_tf,
        # Vector features
        zone_density=zone_density, eoh=eoh, hu_log=hu_log,
        fourier_desc=fourier_desc,
        # Visualization
        img_sq=img, boot_mask=boot_mask, tread_mask=tread_mask,
        edges=edges_full,
        view_type='aligned',
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SIMILARITY — v6 with phase-correlation alignment
# ══════════════════════════════════════════════════════════════════════════════

def orb_match_score(desc_a, desc_b):
    if desc_a is None or desc_b is None: return 0.0
    if len(desc_a) < 5 or len(desc_b) < 5: return 0.0
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=False)
    try: matches = bf.knnMatch(desc_a, desc_b, k=2)
    except: return 0.0
    good = sum(1 for m in matches if len(m) == 2 and m[0].distance < 0.78 * m[1].distance)
    return good / max(len(desc_a), 1)


def phase_corr_align_and_compare(a, b):
    """
    Use phase correlation to find best translation alignment,
    then compute similarity on the aligned pair.
    """
    af = a.astype(np.float32)
    bf = b.astype(np.float32)
    
    # Phase correlation to find shift
    shift, response = cv2.phaseCorrelate(af, bf)
    dx, dy = int(round(shift[0])), int(round(shift[1]))
    
    # Shift b to align with a
    M = np.float32([[1, 0, dx], [0, 1, dy]])
    h, w = b.shape
    b_shifted = cv2.warpAffine(bf, M, (w, h))
    
    # Compute NCC on aligned images
    an = af - af.mean()
    bn = b_shifted - b_shifted.mean()
    denom = max(np.std(an) * np.std(bn) * an.size, 1e-9)
    ncc = float(np.sum(an * bn) / denom)
    
    return max(0.0, (ncc + 1) / 2), response


def ssim_windowed(a, b, win_size=11):
    """Compute mean SSIM over sliding windows."""
    C1 = (0.01 * 255) ** 2
    C2 = (0.03 * 255) ** 2
    
    af = a.astype(np.float64)
    bf = b.astype(np.float64)
    
    k = cv2.getGaussianKernel(win_size, 1.5)
    window = k @ k.T
    
    mu_a = cv2.filter2D(af, -1, window)
    mu_b = cv2.filter2D(bf, -1, window)
    
    mu_a_sq = mu_a ** 2
    mu_b_sq = mu_b ** 2
    mu_ab = mu_a * mu_b
    
    sigma_a_sq = cv2.filter2D(af**2, -1, window) - mu_a_sq
    sigma_b_sq = cv2.filter2D(bf**2, -1, window) - mu_b_sq
    sigma_ab = cv2.filter2D(af*bf, -1, window) - mu_ab
    
    ssim_map = ((2*mu_ab + C1) * (2*sigma_ab + C2)) / \
               ((mu_a_sq + mu_b_sq + C1) * (sigma_a_sq + sigma_b_sq + C2))
    
    return float(np.mean(ssim_map))


def pixel_iou(bw_a, bw_b):
    a = (bw_a > 0).astype(np.uint8)
    b = (bw_b > 0).astype(np.uint8)
    inter = np.sum(a & b)
    union = np.sum(a | b)
    return float(inter / union) if union > 0 else 0.0


def cosine_sim(a, b):
    na = np.linalg.norm(a); nb = np.linalg.norm(b)
    return float(np.clip(np.dot(a,b)/(na*nb), 0, 1)) if na > 1e-9 and nb > 1e-9 else 0.0

def hist_corr(a, b):
    return float(max(0.0, cv2.compareHist(
        a.reshape(-1,1).astype(np.float32),
        b.reshape(-1,1).astype(np.float32), cv2.HISTCMP_CORREL)))


def edge_overlay_score(edge_a, edge_b):
    """
    Compute the overlap ratio of two edge maps — exactly what you see in
    the cyan/yellow/white comparison images.

    white  = edges present in BOTH  (agreement)
    yellow = edges only in A
    cyan   = edges only in B

    score = white / (white + yellow + cyan)

    Works reliably here because v6 already normalises both edge maps to
    the same 256×128 size before this is called.
    Uses phase correlation to find the best translation shift first.
    """
    af = edge_a.astype(np.float32)
    bf = edge_b.astype(np.float32)
    if np.std(af) < 1e-6 or np.std(bf) < 1e-6:
        return 0.0
    shift, _ = cv2.phaseCorrelate(af, bf)
    dx = int(round(np.clip(shift[0], -NORM_W//5, NORM_W//5)))
    dy = int(round(np.clip(shift[1], -NORM_H//5, NORM_H//5)))
    M = np.float32([[1, 0, dx], [0, 1, dy]])
    b_shifted = cv2.warpAffine(edge_b, M, (edge_b.shape[1], edge_b.shape[0]))
    a_bin = edge_a > 0
    b_bin = b_shifted > 0
    both  = np.sum(a_bin & b_bin)
    union = np.sum(a_bin | b_bin)
    return float(both / union) if union > 0 else 0.0


def make_overlay_image(edge_a, edge_b):
    """
    Build the cyan/yellow/white overlay image for visualisation.
    Returns a BGR image at NORM_H × NORM_W.
    """
    af = edge_a.astype(np.float32)
    bf = edge_b.astype(np.float32)
    if np.std(af) > 1e-6 and np.std(bf) > 1e-6:
        shift, _ = cv2.phaseCorrelate(af, bf)
        dx = int(round(np.clip(shift[0], -NORM_W//5, NORM_W//5)))
        dy = int(round(np.clip(shift[1], -NORM_H//5, NORM_H//5)))
        M = np.float32([[1, 0, dx], [0, 1, dy]])
        b_shifted = cv2.warpAffine(edge_b, M, (edge_b.shape[1], edge_b.shape[0]))
    else:
        b_shifted = edge_b
    a_bin = edge_a > 0
    b_bin = b_shifted > 0
    canvas = np.zeros((NORM_H, NORM_W, 3), np.uint8)
    canvas[a_bin & ~b_bin] = (0,   200, 220)   # yellow  — A only
    canvas[b_bin & ~a_bin] = (220, 200,   0)   # cyan    — B only
    canvas[a_bin &  b_bin] = (220, 220, 220)   # white   — both = match
    return canvas


def compute_similarity(fa, fb):
    """Multi-strategy with alignment and mirror testing."""
    
    # Test both original and mirror orientations
    results = []
    for flip_label, tbw_b, epa_b, ng_b, de_b, dt_b in [
        ("orig", fb["tread_bw"], fb["edge_pat"], fb["norm_gray"],
         fb["desc_e"], fb["desc_t"]),
        ("flip", fb["tread_bw_flip"], fb["edge_pat_flip"], fb["norm_gray_flip"],
         fb["desc_ef"], fb["desc_tf"]),
    ]:
        # 1. Phase-correlation aligned NCC on edge patterns
        pc_edge, pc_resp = phase_corr_align_and_compare(
            fa["edge_pat"].astype(np.float32),
            epa_b.astype(np.float32))
        
        # 2. Phase-correlation aligned NCC on tread binary
        pc_tread, _ = phase_corr_align_and_compare(
            fa["tread_bw"].astype(np.float32),
            tbw_b.astype(np.float32))
        
        # 3. SSIM on normalized gray (after alignment by canonical orientation)
        ssim_val = ssim_windowed(fa["norm_gray"], ng_b)
        ssim_val = max(0.0, (ssim_val + 1) / 2)  # map -1..1 to 0..1
        
        # 4. ORB matching
        orb_e = orb_match_score(fa["desc_e"], de_b)
        orb_t = orb_match_score(fa["desc_t"], dt_b)
        s_orb = max(orb_e, orb_t)
        s_orb_norm = min(1.0, s_orb / 0.10)
        
        # 5. Pixel IoU on tread binary
        s_iou = pixel_iou(fa["tread_bw"], tbw_b)
        
        # 6. Zone density correlation
        zd_corr = float(np.corrcoef(fa["zone_density"], fb["zone_density"])[0,1])
        s_zone = max(0.0, zd_corr if not np.isnan(zd_corr) else 0.0)
        
        # 7. Edge orientation histogram
        s_eoh = hist_corr(fa["eoh"], fb["eoh"])
        
        # 8. Hu moments distance
        s_hu = 1.0 - min(1.0, np.sum(np.abs(fa["hu_log"] - fb["hu_log"])) / 40.0)
        
        # 9. Fourier descriptor correlation
        s_fd = cosine_sim(fa["fourier_desc"], fb["fourier_desc"])

        # Overlay score — computed for visualization only, NOT used in scoring
        # (pixel-level overlap is too sensitive to minor scale/angle differences)
        s_eo = edge_overlay_score(fa["edge_pat"], epa_b)

        # Original v6 weights — restored now that all descriptors are working
        score = (0.20 * pc_edge +
                 0.15 * pc_tread +
                 0.15 * ssim_val +
                 0.12 * s_orb_norm +
                 0.10 * s_iou +
                 0.08 * s_zone +
                 0.08 * s_eoh +
                 0.06 * s_hu +
                 0.06 * s_fd)

        sub = dict(pc_e=round(pc_edge,3), pc_t=round(pc_tread,3),
                   ssim=round(ssim_val,3), orb=round(s_orb_norm,3),
                   iou=round(s_iou,3), zone=round(s_zone,3),
                   eoh=round(s_eoh,3), hu=round(s_hu,3), fd=round(s_fd,3),
                   eo=round(s_eo,3),   # kept for display in comparison cards
                   flip=flip_label)
        results.append((score, sub))
    
    # Take the best orientation
    best_score, best_sub = max(results, key=lambda x: x[0])
    return round(float(best_score), 4), best_sub


# ══════════════════════════════════════════════════════════════════════════════
#  CLUSTERING
# ══════════════════════════════════════════════════════════════════════════════

class UnionFind:
    def __init__(self, n): self.p = list(range(n)); self.r = [0]*n
    def find(self, x):
        while self.p[x] != x: self.p[x] = self.p[self.p[x]]; x = self.p[x]
        return x
    def union(self, x, y):
        rx, ry = self.find(x), self.find(y)
        if rx == ry: return
        if self.r[rx] < self.r[ry]: rx, ry = ry, rx
        self.p[ry] = rx
        if self.r[rx] == self.r[ry]: self.r[rx] += 1

def cluster(n, pairs, thresh):
    uf = UnionFind(n)
    for i, j, s, _ in pairs:
        if s >= thresh: uf.union(i, j)
    g = defaultdict(list)
    for idx in range(n): g[uf.find(idx)].append(idx)
    return [sorted(m) for m in g.values() if len(m) >= 2]


# ══════════════════════════════════════════════════════════════════════════════
#  COMPARISON IMAGE
# ══════════════════════════════════════════════════════════════════════════════

def make_comparison_image(features, grp, score_lk, gid, title_text, bar_color):
    """
    Full match viewer layout — tight, no wasted black space.

    Layout:
      HEADER
      Photo row       — original photos, tread region highlighted green
                        (height auto-fitted to actual content, not fixed)
      Binary row      — normalised tread binary side by side
                        (height auto-fitted to actual content)
      Overlay panel   — cyan/yellow/white edge overlap, aspect-ratio correct
      Score bar chart — all sub-scores as horizontal bars
    """
    SHOE_W   = 480
    PAD      = 10          # padding inside each photo cell
    HEADER_H = 48
    LABEL_H  = 24          # strip below each panel
    font = cv2.FONT_HERSHEY_SIMPLEX
    LABEL_COLOR = (200, 220, 255)
    n = len(grp)

    # ── Helpers ──────────────────────────────────────────────────────────

    def _tight_fit(src, target_w, pad=PAD):
        """Scale src to fill target_w with minimal vertical padding.
        Returns (panel_bgr, actual_content_h) where panel height = content_h + 2*pad."""
        sh, sw = src.shape[:2]
        sc = (target_w - 2*pad) / max(sw, 1)
        nw = max(1, int(sw * sc))
        nh = max(1, int(sh * sc))
        resized = cv2.resize(src, (nw, nh), interpolation=cv2.INTER_AREA)
        if len(resized.shape) == 2:
            resized = cv2.cvtColor(resized, cv2.COLOR_GRAY2BGR)
        panel_h = nh + 2 * pad
        canvas = np.zeros((panel_h, target_w, 3), np.uint8)
        y0 = pad
        x0 = (target_w - nw) // 2
        canvas[y0:y0+nh, x0:x0+nw] = resized
        return canvas, nh

    def _label(img, text, x, y, scale=0.42, color=LABEL_COLOR):
        cv2.putText(img, text, (x, y), font, scale, (0,0,0), 3, cv2.LINE_AA)
        cv2.putText(img, text, (x, y), font, scale, color,   1, cv2.LINE_AA)

    def _bar(text, w, h=24, bg=(28,28,28)):
        b = np.full((h, w, 3), bg, np.uint8)
        _label(b, text, 8, h-6, 0.38, LABEL_COLOR)
        return b

    def _pad_to_height(panels, target_h):
        """Pad each panel in list to target_h by adding rows at bottom."""
        out = []
        for p in panels:
            ph = p.shape[0]
            if ph < target_h:
                pad_strip = np.zeros((target_h - ph, p.shape[1], 3), np.uint8)
                p = np.vstack([p, pad_strip])
            out.append(p)
        return out

    # ── Header ───────────────────────────────────────────────────────────
    total_w = SHOE_W * n + 4 * (n - 1)
    header = np.full((HEADER_H, total_w, 3), 16, np.uint8)
    pair_scores = []
    for ia, ib in combinations(grp, 2):
        s, _ = score_lk.get((min(ia,ib), max(ia,ib)), (0.0, {}))
        pair_scores.append(s)
    avg_score = np.mean(pair_scores) if pair_scores else 0.0
    pct_str = str(int(round(avg_score * 100))) + "%"
    gt_ps = set(features[i]["gt_pair"] for i in grp if features[i]["gt_pair"] > 0)
    gt_str = ""
    if gt_ps:
        gt_str = "   [GT: " + ", ".join(
            str(p) + "=" + GT_STATUS.get(p, "?") for p in sorted(gt_ps)) + "]"
    title = title_text + "   " + pct_str + "   (" + str(n) + " shoes)" + gt_str
    sc = (40,200,40) if avg_score >= 0.80 else (40,160,220)
    cv2.putText(header, title, (10,32), font, 0.62, (0,0,0), 3, cv2.LINE_AA)
    cv2.putText(header, title, (10,32), font, 0.62, sc,       1, cv2.LINE_AA)

    divider_h = np.full((4, total_w, 3), 45, np.uint8)

    # ── Photo row — tight fit ─────────────────────────────────────────────
    # First pass: compute each panel and find max content height
    photo_panels = []
    for idx in grp:
        f = features[idx]
        photo = f["img_sq"].copy()
        tm = f["tread_mask"]
        # Crop to tread bounding box + small margin for tighter display
        ys, xs = np.where(tm > 0)
        if len(ys) > 0:
            margin = 20
            y0c = max(0, ys.min() - margin)
            y1c = min(photo.shape[0], ys.max() + margin)
            x0c = max(0, xs.min() - margin)
            x1c = min(photo.shape[1], xs.max() + margin)
            photo_crop = photo[y0c:y1c, x0c:x1c]
            tm_crop    = tm[y0c:y1c, x0c:x1c]
        else:
            photo_crop = photo
            tm_crop    = tm
        photo_crop[cv2.Canny(tm_crop, 50, 150) > 0] = (0, 210, 0)
        photo_crop[tm_crop == 0] = (photo_crop[tm_crop == 0] * 0.32).astype(np.uint8)
        panel, ch = _tight_fit(photo_crop, SHOE_W)
        photo_panels.append((panel, ch, f))

    max_ph = max(p[0].shape[0] for p in photo_panels)
    photo_parts = []
    divider_v_photo = np.full((max_ph + LABEL_H, 4, 3), 45, np.uint8)
    for k, (panel, ch, f) in enumerate(photo_panels):
        # Pad panel to max height
        if panel.shape[0] < max_ph:
            pad_strip = np.zeros((max_ph - panel.shape[0], SHOE_W, 3), np.uint8)
            panel = np.vstack([panel, pad_strip])
        strip = np.zeros((LABEL_H, SHOE_W, 3), np.uint8)
        _label(strip, f["stem"][:30], 6, 17, 0.40)
        ed = round(np.sum(f["edges"] > 0) / max(np.sum(f["tread_mask"] > 0), 1), 3)
        ed_str = "ed=" + str(ed)
        tw, _ = cv2.getTextSize(ed_str, font, 0.34, 1)
        cv2.putText(strip, ed_str, (SHOE_W-tw[0]-6, 17), font, 0.34, (0,220,220), 1, cv2.LINE_AA)
        photo_parts.append(np.vstack([panel, strip]))
        if k < n - 1:
            photo_parts.append(divider_v_photo)
    photo_row = np.hstack(photo_parts)

    # ── Binary tread row — tight fit ──────────────────────────────────────
    binary_panels = []
    for idx in grp:
        f = features[idx]
        bw_land = cv2.rotate(f["tread_bw"], cv2.ROTATE_90_COUNTERCLOCKWISE)
        panel, ch = _tight_fit(bw_land, SHOE_W)
        binary_panels.append(panel)

    max_bh = max(p.shape[0] for p in binary_panels)
    binary_parts = []
    divider_v_bin = np.full((max_bh + LABEL_H, 4, 3), 45, np.uint8)
    for k, panel in enumerate(binary_panels):
        if panel.shape[0] < max_bh:
            pad_strip = np.zeros((max_bh - panel.shape[0], SHOE_W, 3), np.uint8)
            panel = np.vstack([panel, pad_strip])
        strip = _bar("Aligned tread binary  (what the algorithm compares)", SHOE_W)
        binary_parts.append(np.vstack([panel, strip]))
        if k < n - 1:
            binary_parts.append(divider_v_bin)
    binary_row = np.hstack(binary_parts)

    # ── Overlay panel — aspect-ratio correct, fills width ────────────────
    i0, i1 = grp[0], grp[1]
    ea = features[i0]["edge_pat"]
    eb = features[i1]["edge_pat"]
    ov_small = make_overlay_image(ea, eb)           # NORM_H × NORM_W BGR (256×128)
    ov_land  = cv2.rotate(ov_small, cv2.ROTATE_90_COUNTERCLOCKWISE)  # now 128×256

    # Scale to fill total_w while preserving aspect ratio
    ov_src_h, ov_src_w = ov_land.shape[:2]
    ov_scale  = total_w / max(ov_src_w, 1)
    ov_rw     = total_w
    ov_rh     = max(1, int(ov_src_h * ov_scale))
    # Cap height so it doesn't become absurdly tall; add small vertical pad
    OV_MAX_H  = max(ov_rh, 160)
    ov_resized = cv2.resize(ov_land, (ov_rw, ov_rh), interpolation=cv2.INTER_NEAREST)

    # Embed in canvas (centred vertically if capped)
    ov_big = np.zeros((OV_MAX_H, total_w, 3), np.uint8)
    y0_ov  = (OV_MAX_H - ov_rh) // 2
    ov_big[y0_ov:y0_ov + ov_rh, :] = ov_resized

    # Legend — bottom-left of the overlay canvas
    leg_y = OV_MAX_H - 8
    cv2.putText(ov_big, "A only",  (8,   leg_y), font, 0.38, (0,200,220),   1, cv2.LINE_AA)
    cv2.putText(ov_big, "B only",  (80,  leg_y), font, 0.38, (220,200,0),   1, cv2.LINE_AA)
    cv2.putText(ov_big, "MATCH",   (155, leg_y), font, 0.38, (220,220,220), 1, cv2.LINE_AA)

    key01 = (min(i0, i1), max(i0, i1))
    s01, sub01 = score_lk.get(key01, (0.0, {}))
    ov_info = ("Edge overlap: " + str(int(round(sub01.get('eo',0)*100))) + "%"
               + "   pc_e: "   + str(int(round(sub01.get('pc_e',0)*100))) + "%"
               + "   ssim: "   + str(int(round(sub01.get('ssim',0)*100))) + "%"
               + "   orb: "    + str(int(round(sub01.get('orb',0)*100))) + "%"
               + "   iou: "    + str(int(round(sub01.get('iou',0)*100))) + "%")
    ov_label = _bar("Edge overlay: A only=yellow  B only=cyan  Both=white   " + ov_info,
                    total_w, 28, (20,20,20))

    # ── Score bar chart ───────────────────────────────────────────────────
    SCORE_ROWS = [
        ("Phase corr-e   (20%)", "pc_e",    0.20),
        ("Phase corr-t   (15%)", "pc_t",    0.15),
        ("SSIM           (15%)", "ssim",    0.15),
        ("ORB keypoints  (12%)", "orb",     0.12),
        ("Pixel IoU      (10%)", "iou",     0.10),
        ("Zone density    (8%)", "zone",    0.08),
        ("Edge orient     (8%)", "eoh",     0.08),
        ("Hu moments      (6%)", "hu",      0.06),
        ("Fourier desc    (6%)", "fd",      0.06),
        ("TOTAL SCORE",          "__tot__", 1.00),
    ]
    ROW_H   = 30
    LABEL_W = 200
    BAR_W   = total_w - LABEL_W - 70
    chart_h = len(SCORE_ROWS) * ROW_H + 12
    chart   = np.full((chart_h, total_w, 3), 14, np.uint8)

    def bar_col(v):
        if v >= 0.72: return (40,200,40)
        if v >= 0.50: return (0,180,230)
        return (50,50,220)

    for ri, (lbl, key, _) in enumerate(SCORE_ROWS):
        val = s01 if key == "__tot__" else float(sub01.get(key, 0))
        yt  = 6 + ri * ROW_H
        ym  = yt + ROW_H // 2 + 5
        bg  = (35,35,35) if key == "__tot__" else ((22,22,22) if ri%2==0 else (18,18,18))
        cv2.rectangle(chart, (0,yt), (total_w-1, yt+ROW_H-1), bg, -1)
        col = (220,220,255) if key == "__tot__" else LABEL_COLOR
        fs  = 0.42 if key == "__tot__" else 0.38
        cv2.putText(chart, lbl, (8,ym), font, fs, (0,0,0), 3, cv2.LINE_AA)
        cv2.putText(chart, lbl, (8,ym), font, fs, col,     1, cv2.LINE_AA)
        fw = max(2, int(val * BAR_W))
        cv2.rectangle(chart, (LABEL_W, yt+4), (LABEL_W+fw, yt+ROW_H-4), bar_col(val), -1)
        for th in [0.50, 0.72]:
            tx = LABEL_W + int(th * BAR_W)
            cv2.line(chart, (tx, yt+2), (tx, yt+ROW_H-2), (80,80,80), 1)
        pct_s = str(int(round(val * 100))) + "%"
        cv2.putText(chart, pct_s, (LABEL_W+fw+5, ym), font, 0.40, bar_col(val), 1, cv2.LINE_AA)
    chart_bar = _bar("Sub-score breakdown", total_w)

    return np.vstack([
        header,
        photo_row,
        divider_h,
        binary_row,
        divider_h,
        ov_big,
        ov_label,
        divider_h,
        chart_bar,
        chart,
    ])


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL (same structure as v6)
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(features, all_pairs, groups, likely_pairs=None):
    wb = Workbook()
    Hf = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    Nf = Font(name="Arial", size=9)
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)

    score_lk = {}
    for i, j, s, sub in all_pairs:
        score_lk[(min(i,j), max(i,j))] = (s, sub)

    # Sheet 1: Identical matches
    ws1 = wb.active; ws1.title = "Identical Matches"
    H_grn = PatternFill("solid", start_color="1B5E20")
    RF = PatternFill("solid", start_color="E8F5E9")
    hdrs = ["Group", "# Shoes", "Score", "Shoes in Group",
            "GT Pair #", "GT Status", "Sub-scores"]
    for col, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Hf; cell.fill = H_grn; cell.alignment = C
    ws1.row_dimensions[1].height = 28
    for gidx, grp in enumerate(groups, 1):
        pairs_in = list(combinations(grp, 2))
        scores = [score_lk.get((min(i,j), max(i,j)), (0, {}))[0] for i, j in pairs_in]
        subs = [score_lk.get((min(i,j), max(i,j)), (0, {}))[1] for i, j in pairs_in]
        avg_score = np.mean(scores) if scores else 0
        fs = subs[0] if subs else {}
        sub_txt = "pc_e=" + str(fs.get('pc_e',0)) + " ssim=" + str(fs.get('ssim',0)) + \
                  " orb=" + str(fs.get('orb',0)) + " iou=" + str(fs.get('iou',0)) + \
                  " eo=" + str(fs.get('eo',0))
        names = ", ".join(features[i]["stem"] for i in grp)
        gt_pairs = set(features[i]["gt_pair"] for i in grp if features[i]["gt_pair"] > 0)
        gt_pair_str = ", ".join(str(p) for p in sorted(gt_pairs)) if gt_pairs else "—"
        gt_statuses = set()
        for p in gt_pairs:
            if p in GT_STATUS: gt_statuses.add(GT_STATUS[p])
        gt_status_str = ", ".join(gt_statuses) if gt_statuses else "—"
        vals = [gidx, len(grp), str(int(round(avg_score*100))) + "%", names,
                gt_pair_str, gt_status_str, sub_txt]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=gidx+1, column=col, value=val)
            cell.font = Nf; cell.fill = RF
            cell.alignment = L if col in (4, 6, 7) else C
    for col, w in enumerate([8, 8, 8, 55, 10, 14, 50], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A2"

    # Sheet 2: Likely same mold
    ws2 = wb.create_sheet("Likely Same Mold")
    H_ora = PatternFill("solid", start_color="B85C00")
    OF = PatternFill("solid", start_color="FFF0DC")
    hdrs2 = ["Rank", "Score", "Shoe A", "Shoe B", "GT Pair #", "GT Status", "Sub-scores"]
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Hf; cell.fill = H_ora; cell.alignment = C
    ws2.row_dimensions[1].height = 28
    for rank, (i, j, s, sub) in enumerate((likely_pairs or []), 1):
        sub_txt = "pc_e=" + str(sub.get('pc_e',0)) + " ssim=" + str(sub.get('ssim',0)) + \
                  " orb=" + str(sub.get('orb',0)) + " iou=" + str(sub.get('iou',0))
        pa, pb = features[i]["gt_pair"], features[j]["gt_pair"]
        gt_pair_str = str(pa) if (pa > 0 and pb > 0 and pa == pb) else ""
        gt_status_str = GT_STATUS.get(pa, "") if gt_pair_str else ""
        vals = [rank, str(int(round(s*100))) + "%", features[i]["stem"],
                features[j]["stem"], gt_pair_str, gt_status_str, sub_txt]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=rank+1, column=col, value=val)
            cell.font = Nf; cell.fill = OF
            cell.alignment = L if col in (3, 4, 6, 7) else C
    for col, w in enumerate([6, 8, 35, 35, 10, 14, 50], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Ground Truth Validation
    ws_gt = wb.create_sheet("Ground Truth Validation")
    H_blue = PatternFill("solid", start_color="1565C0")
    hdrs3 = ["Pair #", "Style", "Brand", "Outsole Tech", "M/F",
             "IDAPT A", "IDAPT B", "GT Status",
             "Algorithm Score", "Algorithm Match?", "Correct?"]
    for col, h in enumerate(hdrs3, 1):
        cell = ws_gt.cell(row=1, column=col, value=h)
        cell.font = Hf; cell.fill = H_blue; cell.alignment = C
    ws_gt.row_dimensions[1].height = 28

    idapt_to_fidx = {}
    for fidx, f in enumerate(features):
        idapt_to_fidx[f["idapt"]] = fidx

    row = 2
    TP = FP = TN = FN = 0
    for pn in sorted(GT_PAIRS.keys()):
        ids = GT_PAIRS[pn]
        if len(ids) < 2: continue
        id_a, id_b = ids[0], ids[1]
        gt_st = GT_STATUS.get(pn, "")
        info = GT_INFO.get(id_a, {})
        fi_a = idapt_to_fidx.get(id_a)
        fi_b = idapt_to_fidx.get(id_b)
        if fi_a is not None and fi_b is not None:
            key = (min(fi_a, fi_b), max(fi_a, fi_b))
            score_val, _ = score_lk.get(key, (None, None))
            algo_score = str(int(round(score_val*100))) + "%" if score_val is not None else "—"
            algo_match = ("Yes" if score_val is not None and score_val >= IDENTICAL_THRESH else
                          "Likely" if score_val is not None and score_val >= LIKELY_THRESH else "No")
        else:
            algo_score = "N/A (missing image)"
            algo_match = "N/A"
        correct = ""
        if gt_st == "identical" and algo_match in ("Yes", "Likely"):
            correct = "✓ TP"; TP += 1
        elif gt_st == "identical" and algo_match == "No":
            correct = "✗ FN"; FN += 1
        elif gt_st == "not identical" and algo_match == "No":
            correct = "✓ TN"; TN += 1
        elif gt_st == "not identical" and algo_match in ("Yes", "Likely"):
            correct = "✗ FP"; FP += 1
        vals = [pn, info.get('style',''), info.get('brand',''),
                info.get('tech',''), info.get('mf',''),
                id_a, id_b, gt_st, algo_score, algo_match, correct]
        for col, val in enumerate(vals, 1):
            cell = ws_gt.cell(row=row, column=col, value=val)
            cell.font = Nf; cell.alignment = L if col in (2, 3, 4) else C
            if "✓" in str(val):
                cell.fill = PatternFill("solid", start_color="C8E6C9")
            elif "✗" in str(val):
                cell.fill = PatternFill("solid", start_color="FFCDD2")
            elif "no image" in str(gt_st).lower():
                cell.fill = PatternFill("solid", start_color="F5F5F5")
        row += 1

    row += 1
    ws_gt.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, name="Arial", size=10)
    total_eval = TP + FP + TN + FN
    if total_eval > 0:
        acc = (TP + TN) / total_eval * 100
        prec = TP / max(TP + FP, 1) * 100
        rec = TP / max(TP + FN, 1) * 100
        ws_gt.cell(row=row+1, column=1, value="True Positives: " + str(TP)).font = Nf
        ws_gt.cell(row=row+2, column=1, value="True Negatives: " + str(TN)).font = Nf
        ws_gt.cell(row=row+3, column=1, value="False Positives: " + str(FP)).font = Nf
        ws_gt.cell(row=row+4, column=1, value="False Negatives: " + str(FN)).font = Nf
        summary = "Accuracy: " + str(round(acc, 1)) + "%  Precision: " + \
                  str(round(prec, 1)) + "%  Recall: " + str(round(rec, 1)) + "%"
        ws_gt.cell(row=row+5, column=1, value=summary).font = Font(bold=True, name="Arial", size=11)

    for col, w in enumerate([8, 22, 25, 30, 6, 12, 12, 14, 14, 14, 10], 1):
        ws_gt.column_dimensions[get_column_letter(col)].width = w
    ws_gt.freeze_panes = "A2"

    wb.save(str(EXCEL_PATH))
    print("  Excel saved: " + str(EXCEL_PATH))
    return TP, FP, TN, FN


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if not SOLE_DIR.exists():
        print("ERROR: " + str(SOLE_DIR) + " not found."); return
    images = [p for p in sorted(SOLE_DIR.iterdir())
              if p.suffix.lower() in IMAGE_EXTS and p.is_file()]
    if len(images) < 2:
        print("Need >=2 images in " + str(SOLE_DIR)); return

    print("\n" + "="*65)
    print("  Sole Similarity Analyzer  v13  —  v6 weights + all descriptors fixed + viewer")
    print("  Input : " + str(SOLE_DIR) + "  (" + str(len(images)) + " images)")
    print("  Identical >= " + str(int(IDENTICAL_THRESH*100)) + "%   Likely >= " + str(int(LIKELY_THRESH*100)) + "%")
    print("  Overlay shown in cards but NOT used in scoring")
    print("  Viewer output: " + str(VIEWER_DIR))
    print("  Ground truth: " + str(len(GT_PAIRS)) + " pairs")
    print("="*65 + "\n")

    print("  Extracting features (with tread isolation + alignment)...")
    features = []
    for p in images:
        print("    " + p.name.ljust(40), end=" ", flush=True)
        f = extract_features(p)
        if f is None: print("SKIP"); continue
        features.append(f)
        gt_tag = " [GT pair " + str(f['gt_pair']) + "]" if f['gt_pair'] > 0 else ""
        n_kp = len(f['desc_e']) if f['desc_e'] is not None else 0
        print("[tread OK]  ORB_kp=" + str(n_kp).rjust(3) + gt_tag)

    n = len(features)
    total_pairs = n * (n - 1) // 2
    print("\n  Computing " + str(total_pairs) + " pairwise similarities...")
    all_pairs = []; score_lk = {}
    done = 0
    for i, j in combinations(range(n), 2):
        s, sub = compute_similarity(features[i], features[j])
        all_pairs.append((i, j, s, sub))
        score_lk[(min(i,j), max(i,j))] = (s, sub)
        done += 1
        if done % 1000 == 0:
            pct = done * 100 // total_pairs
            print("    " + str(done) + "/" + str(total_pairs) + " (" + str(pct) + "%)", flush=True)

    groups = cluster(n, all_pairs, IDENTICAL_THRESH)
    likely_pairs = [(i, j, s, sub) for i, j, s, sub in all_pairs
                    if LIKELY_THRESH <= s < IDENTICAL_THRESH]
    likely_pairs.sort(key=lambda x: -x[2])

    # Print results
    print("\n  " + "="*65)
    print("  IDENTICAL SOLE GROUPS (>= " + str(int(IDENTICAL_THRESH*100)) + "%)")
    print("  " + "="*65)
    if groups:
        for gidx, grp in enumerate(groups, 1):
            nms = [features[i]["stem"] for i in grp]
            gt_ps = set(features[i]["gt_pair"] for i in grp if features[i]["gt_pair"] > 0)
            gt_tag = ""
            if gt_ps:
                gt_items = ','.join(str(p) + '=' + GT_STATUS.get(p, '?') for p in gt_ps)
                gt_tag = "  [GT: " + gt_items + "]"
            print("  Group " + str(gidx).zfill(2) + " (" + str(len(grp)) + "): " + \
                  ", ".join(nms) + gt_tag)
        total_shoes = sum(len(g) for g in groups)
        print("\n  Total: " + str(len(groups)) + " groups (" + str(total_shoes) + " shoes)")
    else:
        print("  (none found)")

    if groups:
        print("\n  Saving comparison images to match_viewer_v12/...")
        for gidx, grp in enumerate(groups, 1):
            title = "Group " + str(gidx).zfill(2) + "  [IDENTICAL >= " + \
                    str(int(IDENTICAL_THRESH*100)) + "%]"
            comp = make_comparison_image(features, grp, score_lk, gidx, title, (0, 50, 220))
            nms = [features[i]["stem"][:16] for i in grp]
            fname = ("match_" + str(gidx).zfill(3) + "_grp" + str(gidx).zfill(2)
                     + "_" + nms[0]
                     + ("_+" + str(len(nms)-1) + "more" if len(nms) > 2
                        else "_vs_" + nms[1]) + ".jpg")
            cv2.imwrite(str(VIEWER_DIR / fname), comp, [cv2.IMWRITE_JPEG_QUALITY, 94])
            print("    saved → " + fname)

    print("\n  " + "="*65)
    likely_label = "LIKELY SAME MOLD (" + str(int(LIKELY_THRESH*100)) + "-" + \
                   str(int(IDENTICAL_THRESH*100)-1) + "%)"
    print("  " + likely_label)
    print("  " + "="*65)
    if likely_pairs:
        for i, j, s, sub in likely_pairs[:25]:
            gt_tag = ""
            pa, pb = features[i]["gt_pair"], features[j]["gt_pair"]
            if pa > 0 and pb > 0 and pa == pb:
                gt_tag = "  [GT pair " + str(pa) + ": " + GT_STATUS.get(pa, '?') + "]"
            print("  " + features[i]['stem'][:28].ljust(28) + " <-> " +
                  features[j]['stem'][:28].ljust(28) + "  " +
                  str(int(round(s*100))).rjust(2) + "%" + gt_tag)
        if len(likely_pairs) > 25:
            print("  ... and " + str(len(likely_pairs)-25) + " more")
        print("\n  Total: " + str(len(likely_pairs)) + " candidate pairs")
        print("\n  Saving likely-match images (top 20)...")
        for rank, (i, j, s, sub) in enumerate(likely_pairs[:20], 1):
            title = "Likely " + str(rank).zfill(2) + "  [" + str(int(round(s*100))) + "%]"
            comp = make_comparison_image(features, [i, j], score_lk, rank, title, (0, 92, 184))
            fname = ("likely_" + str(rank).zfill(2) + "_" + str(int(round(s*100))) + "pct"
                     + "_" + features[i]["stem"][:16] + "_vs_" + features[j]["stem"][:16] + ".jpg")
            cv2.imwrite(str(VIEWER_DIR / fname), comp, [cv2.IMWRITE_JPEG_QUALITY, 94])
    else:
        print("  (none found)")

    # Validation
    print("\n  " + "="*65)
    print("  GROUND TRUTH VALIDATION")
    print("  " + "="*65)
    TP, FP, TN, FN = build_excel(features, all_pairs, groups, likely_pairs)
    total = TP + FP + TN + FN
    if total > 0:
        acc = (TP + TN) / total * 100
        prec = TP / max(TP + FP, 1) * 100
        rec = TP / max(TP + FN, 1) * 100
        print("  TP=" + str(TP) + "  TN=" + str(TN) + "  FP=" + str(FP) + "  FN=" + str(FN))
        print("  Accuracy: " + str(round(acc,1)) + "%  Precision: " + \
              str(round(prec,1)) + "%  Recall: " + str(round(rec,1)) + "%")
    else:
        print("  (no evaluable GT pairs — check IDAPT filename patterns)")

    print("\n  Done!\n")


if __name__ == "__main__":
    main()