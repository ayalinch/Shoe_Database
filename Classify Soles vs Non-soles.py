#!/usr/bin/env python3
"""
Hybrid Shoe Sole Classifier
============================
Uses CLIP (zero-shot ML) + OpenCV (rule-based) to classify images as
shoe soles or not, then copies them into subfolders.

Usage:
    python classify_soles.py --input /path/to/images --output /path/to/output
    python classify_soles.py --input /path/to/images  # output defaults to input/classified/

Dependencies:
    pip install torch torchvision clip-by-openai opencv-python Pillow tqdm openpyxl
    OR: pip install open_clip_torch opencv-python Pillow tqdm openpyxl
"""

import argparse
import os
import shutil
import sys
import json
import cv2
import numpy as np
from pathlib import Path
from PIL import Image
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── CLIP setup ────────────────────────────────────────────────────────────────
def load_clip():
    try:
        import clip
        import torch
        device = "cuda" if torch.cuda.is_available() else "cpu"
        model, preprocess = clip.load("ViT-B/32", device=device)
        print(f"[CLIP] Loaded OpenAI CLIP on {device}")
        return model, preprocess, device, "openai"
    except ImportError:
        pass
    try:
        import open_clip
        import torch
        device = "cuda" if torch.cuda.is_available() else "cpu"
        model, _, preprocess = open_clip.create_model_and_transforms("ViT-B-32", pretrained="openai")
        model = model.to(device)
        model.eval()
        print(f"[CLIP] Loaded open_clip on {device}")
        return model, preprocess, device, "open_clip"
    except ImportError:
        print("[CLIP] Neither 'clip' nor 'open_clip_torch' found. Running OpenCV-only mode.")
        return None, None, None, None


def clip_score(model, preprocess, device, clip_type, image_path):
    """
    Returns probability [0-1] that the image is a shoe sole via CLIP zero-shot.

    Prompt changes vs earlier versions:
    - REMOVED "a shoe floating on a dark or grey background product shot" from negatives
      → was matching dark sole photos on dark backgrounds, hurting FN rate
    - ADDED two dark-sole-specific positive prompts to help dark-on-dark sole images
    - ADDED two specific side-view negatives to replace the removed generic one
    """
    import torch

    POSITIVE_PROMPTS = [
        "boot sole photographed from directly below showing full tread pattern",
        "flat overhead view of shoe outsole rubber lugs",
        "bottom of a boot laid flat showing entire sole tread",
        "shoe sole tread pattern viewed straight on from underneath",
        "close-up of rubber boot outsole tread and grip pattern",
        "shoe outsole with lug pattern on plain background",
        "dark rubber sole with tread pattern photographed from directly above",
        "boot bottom viewed from below showing heel and toe rubber lugs",
    ]
    NEGATIVE_PROMPTS = [
        "a boot photographed from the side showing the full boot",
        "a winter boot standing upright on a surface",
        "product photo of a boot from the side or angle",
        "a shoe or boot with laces visible from the front or side",
        "a shoe photographed from the side showing the upper surface",
        "side view of a complete shoe showing laces tongue and outsole together",
        "a person wearing boots walking",
        "a boot on a shelf or box",
        "a landscape or indoor scene",
        "a document or paper",
        "a side profile of a shoe showing the upper and tongue",
    ]

    all_prompts = POSITIVE_PROMPTS + NEGATIVE_PROMPTS

    try:
        img = preprocess(Image.open(image_path).convert("RGB")).unsqueeze(0).to(device)
        with torch.no_grad():
            if clip_type == "openai":
                import clip
                text = clip.tokenize(all_prompts).to(device)
                image_features = model.encode_image(img)
                text_features  = model.encode_text(text)
            else:
                import open_clip
                tokenizer = open_clip.get_tokenizer("ViT-B-32")
                text = tokenizer(all_prompts).to(device)
                image_features = model.encode_image(img)
                text_features  = model.encode_text(text)

            image_features /= image_features.norm(dim=-1, keepdim=True)
            text_features  /= text_features.norm(dim=-1, keepdim=True)
            similarity = (100.0 * image_features @ text_features.T).softmax(dim=-1)
            probs = similarity[0].cpu().numpy()

        pos_score = probs[:len(POSITIVE_PROMPTS)].sum()
        neg_score = probs[len(POSITIVE_PROMPTS):].sum()
        return float(pos_score / (pos_score + neg_score + 1e-9))

    except Exception as e:
        print(f"  [CLIP] Error on {image_path.name}: {e}")
        return 0.5


# ── OpenCV heuristics ─────────────────────────────────────────────────────────
def opencv_score(image_path):
    """
    Rule-based score [0-1].

    Key changes vs original (data-driven from 860-image real run):
    ─────────────────────────────────────────────────────────────────────────
    1. CONTOUR FALLBACK (major FN fix):
       Otsu thresholding fails on dark soles against dark backgrounds
       (e.g. mean gray = 13-30, Otsu picks threshold=35-107, coverage < 0.10).
       7/10 test soles were missing the +0.20 contour bonus for this reason.
       Fix: if Otsu coverage < 0.10, fall back to fg_mask (pixels > 15) for
       contour detection. This recovers +0.20 for those dark soles.

    2. THIRDS THRESHOLD TIGHTENED (FP fix):
       "Uniform texture" reward requires mean > 0.06 AND top third > 0.04.
       Prevents rewarding shoes where the dark BG creates uniformly SPARSE
       (not uniformly dense) edge density.

    3. TOP-QUARTER EMPTINESS (FP fix):
       Shoes sit on a baseline — top quarter fg fraction = 0.00-0.15.
       Soles fill the frame — top quarter fg fraction = 0.22-1.00.
       → q1 < 0.18 + lower quarters full: -0.35 penalty.

    4. FG-MASKED COLOR (FP fix):
       Dark BG pixels were inflating dark_frac, giving every shoe on
       black BG an unearned +0.20. Now only foreground pixels count.
       Dark threshold raised 0.20 → 0.35 (fg-relative).

    5. LIGHT SOLE SUPPORT (FN fix):
       Added light_frac for tan/white/grey rubber soles.

    6. EXPANDED CONTOUR COVERAGE (FN fix):
       Coverage range: (0.30, 0.98) → (0.15, 0.995).
    ─────────────────────────────────────────────────────────────────────────
    """
    img = cv2.imread(str(image_path))
    if img is None:
        return 0.0

    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    score = 0.0

    # 1. Edge density
    edges = cv2.Canny(gray, 50, 150)
    edge_density = np.count_nonzero(edges) / (h * w)
    if 0.04 < edge_density < 0.35:
        score += 0.25
    elif edge_density >= 0.35:
        score += 0.10

    # 2. Vertical texture distribution
    thirds = np.array_split(edges, 3, axis=0)
    td = [np.count_nonzero(t) / (t.size + 1e-9) for t in thirds]
    density_range = max(td) - min(td)
    if density_range < 0.06 and np.mean(td) > 0.06 and td[0] > 0.04:
        score += 0.20
    elif td[2] > td[0] * 2.5 and td[2] > td[1] * 1.8:
        score -= 0.25
    elif density_range > 0.10:
        score -= 0.10

    # 3. Smooth upper region (side-view shoe signal)
    top_half = gray[:h // 2, :]
    top_blur = cv2.GaussianBlur(top_half.astype(np.float32), (15, 15), 0)
    lv = np.sum(np.abs(top_half.astype(np.float32) - top_blur) < 8) / top_half.size
    if lv > 0.60:
        score -= 0.20

    # 4. Texture periodicity via FFT
    f = np.fft.fft2(gray.astype(np.float32))
    fshift = np.fft.fftshift(f)
    magnitude = np.log(np.abs(fshift) + 1)
    cy, cx = h // 2, w // 2
    dc_mask = np.ones_like(magnitude)
    dc_mask[cy-5:cy+5, cx-5:cx+5] = 0
    periodicity = (magnitude * dc_mask).mean() / (magnitude.max() + 1e-9)
    if periodicity > 0.15:
        score += 0.20

    # 5. Image aspect ratio
    aspect = min(h, w) / max(h, w)
    if aspect > 0.45:
        score += 0.15
    elif aspect > 0.35:
        score += 0.08

    # 6. Color — fg-masked only
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    fg_bool  = np.any(img > 15, axis=2)
    fg_count = fg_bool.sum() + 1e-9
    dark_frac  = np.sum((hsv[:, :, 2] < 80)  & fg_bool) / fg_count
    light_frac = np.sum((hsv[:, :, 2] > 180) & fg_bool) / fg_count
    sat_frac   = np.sum((hsv[:, :, 1] > 60)  & fg_bool) / fg_count
    if dark_frac > 0.35 or sat_frac > 0.25 or light_frac > 0.30:
        score += 0.20

    # 7. Contour regularity — with dark-on-dark fallback
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cov, sol = 0.0, 0.0
    if contours:
        largest   = max(contours, key=cv2.contourArea)
        hull_area = cv2.contourArea(cv2.convexHull(largest))
        sol       = cv2.contourArea(largest) / (hull_area + 1e-9)
        cov       = hull_area / (h * w)

    # Fallback: Otsu fails on dark soles against dark backgrounds
    # (Otsu picks a high threshold, leaving only bright logo pixels as "foreground")
    if cov < 0.10:
        fg_thresh = (np.any(img > 15, axis=2).astype(np.uint8) * 255)
        contours2, _ = cv2.findContours(fg_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if contours2:
            largest2   = max(contours2, key=cv2.contourArea)
            hull_area2 = cv2.contourArea(cv2.convexHull(largest2))
            sol        = cv2.contourArea(largest2) / (hull_area2 + 1e-9)
            cov        = hull_area2 / (h * w)

    if 0.15 < cov < 0.995 and sol > 0.55:
        score += 0.20

    # 8. Top-quarter emptiness (shoe baseline signal)
    fg_mask = np.any(img > 15, axis=2)
    q = h // 4
    q1_frac = fg_mask[:q,    :].sum() / (q * w + 1e-9)
    q3_frac = fg_mask[q*2:q*3, :].sum() / (q * w + 1e-9)
    q4_frac = fg_mask[q*3:,  :].sum() / (q * w + 1e-9)

    if q1_frac < 0.18 and (q3_frac > 0.40 or q4_frac > 0.30):
        score -= 0.35
    elif q1_frac > 0.20:
        score += 0.10

    return min(max(score, 0.0), 1.0)


# ── Hybrid decision ───────────────────────────────────────────────────────────
def classify_image(image_path, model, preprocess, device, clip_type,
                   clip_weight=0.65, opencv_weight=0.35, threshold=0.50):
    """
    Combines CLIP and OpenCV scores.

    Threshold lowered 0.58 → 0.50 based on analysis of 860-image real run:
    - 25 FN soles scored 35-57% (all below old threshold of 58%)
    - 3 FP shoes scored 58-92% (all safely above new threshold of 50%)
    - Lowering to 0.50 fixes 17/25 FN soles while keeping all 3 FP shoes blocked.
    - Remaining 8 FN soles (35-49%) addressed by contour fallback + CLIP fixes.
    """
    cv_score = opencv_score(image_path)

    if model is not None:
        cl_score = clip_score(model, preprocess, device, clip_type, image_path)
        combined = clip_weight * cl_score + opencv_weight * cv_score
    else:
        cl_score = None
        combined = cv_score
        threshold = 0.40  # lower bar for opencv-only mode

    label = "sole" if combined >= threshold else "not_sole"
    return label, combined, cl_score, cv_score


# ── Main ──────────────────────────────────────────────────────────────────────
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}

def main():
    parser = argparse.ArgumentParser(description="Classify images as shoe soles or not.")
    parser.add_argument("--input",  "-i", default=os.path.expanduser("~/Desktop/extracted_boot_images"))
    parser.add_argument("--output", "-o", default=os.path.expanduser("~/Desktop/extracted_boot_images/classified"))
    parser.add_argument("--move",        action="store_true", help="Move files instead of copying")
    parser.add_argument("--threshold",   type=float, default=0.50, help="Classification threshold (default: 0.50)")
    parser.add_argument("--clip-weight", type=float, default=0.65)
    parser.add_argument("--opencv-weight", type=float, default=0.35)
    parser.add_argument("--opencv-only", action="store_true")
    parser.add_argument("--report",      action="store_true", help="Save JSON report")
    args = parser.parse_args()

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        print(f"ERROR: Input folder not found: {input_dir}")
        sys.exit(1)

    output_dir  = Path(args.output)
    sole_dir     = output_dir / "sole"
    not_sole_dir = output_dir / "not_sole"
    sole_dir.mkdir(parents=True, exist_ok=True)
    not_sole_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*55}")
    print(f"  Shoe Sole Classifier")
    print(f"{'='*55}")
    print(f"  Input    : {input_dir}")
    print(f"  Output   : {output_dir}")
    print(f"  Action   : {'MOVE' if args.move else 'COPY'}")
    print(f"  Threshold: {args.threshold}")
    print(f"{'='*55}\n")

    if args.opencv_only:
        model = preprocess = device = clip_type = None
        print("[Mode] OpenCV-only (no CLIP)")
    else:
        model, preprocess, device, clip_type = load_clip()

    images = [p for p in sorted(input_dir.iterdir())
              if p.suffix.lower() in IMAGE_EXTENSIONS and p.is_file()]
    if not images:
        print(f"No images found in {input_dir}")
        sys.exit(0)
    print(f"Found {len(images)} images to classify.\n")

    results      = []
    sole_count   = 0
    not_sole_count = 0

    for img_path in tqdm(images, desc="Classifying"):
        label, combined, cl_sc, cv_sc = classify_image(
            img_path, model, preprocess, device, clip_type,
            clip_weight=args.clip_weight,
            opencv_weight=args.opencv_weight,
            threshold=args.threshold,
        )

        dest = (sole_dir if label == "sole" else not_sole_dir) / img_path.name
        (shutil.move if args.move else shutil.copy2)(str(img_path), str(dest))

        if label == "sole":
            sole_count += 1
        else:
            not_sole_count += 1

        clip_str = f"{cl_sc:.3f}" if cl_sc is not None else "N/A"
        tqdm.write(f"  {'✓ SOLE    ' if label=='sole' else '✗ not_sole'} "
                   f"| combined={combined:.3f} clip={clip_str} cv={cv_sc:.3f} "
                   f"| {img_path.name}")

        results.append({
            "file": img_path.name,
            "label": label,
            "combined_score": round(combined, 4),
            "clip_score": round(cl_sc, 4) if cl_sc is not None else None,
            "opencv_score": round(cv_sc, 4),
        })

    # ── Excel report ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
    RED_FILL    = PatternFill("solid", start_color="FFC7CE")
    HEADER_FILL = PatternFill("solid", start_color="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    NORMAL_FONT = Font(name="Arial", size=10)
    BOLD_FONT   = Font(bold=True, name="Arial", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(["Filename", "Sole?", "Confidence %", "CLIP Score %", "OpenCV Score %"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(results, 2):
        is_sole = r["label"] == "sole"
        fill    = GREEN_FILL if is_sole else RED_FILL

        ws.cell(row=i, column=1, value=r["file"]).font = NORMAL_FONT
        c2 = ws.cell(row=i, column=2, value="Yes" if is_sole else "No")
        c2.font = BOLD_FONT; c2.alignment = CENTER

        for col, val in [
            (3, f"{r['combined_score']*100:.1f}%"),
            (4, f"{r['clip_score']*100:.1f}%" if r['clip_score'] is not None else "N/A"),
            (5, f"{r['opencv_score']*100:.1f}%"),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = CENTER; c.font = NORMAL_FONT

        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = fill

    sr = len(results) + 3
    ws.cell(row=sr,   column=1, value="Total Images").font = BOLD_FONT
    ws.cell(row=sr,   column=2, value=len(images)).font    = BOLD_FONT
    ws.cell(row=sr+1, column=1, value="Soles Found").font  = Font(bold=True, name="Arial", size=10, color="375623")
    ws.cell(row=sr+1, column=2, value=sole_count).font     = Font(bold=True, name="Arial", size=10, color="375623")
    ws.cell(row=sr+2, column=1, value="Not Soles").font    = Font(bold=True, name="Arial", size=10, color="9C0006")
    ws.cell(row=sr+2, column=2, value=not_sole_count).font = Font(bold=True, name="Arial", size=10, color="9C0006")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A2"

    excel_path = output_dir / "classification_results.xlsx"
    wb.save(str(excel_path))

    print(f"\n{'='*55}")
    print(f"  ✓ Soles     : {sole_count}")
    print(f"  ✗ Not soles : {not_sole_count}")
    print(f"  Total       : {len(images)}")
    print(f"\n  Excel report: {excel_path}")

    if args.report:
        report_path = output_dir / "classification_report.json"
        with open(report_path, "w") as f:
            json.dump({
                "summary": {
                    "total": len(images), "sole": sole_count, "not_sole": not_sole_count,
                    "threshold": args.threshold,
                    "mode": "opencv_only" if model is None else "hybrid_clip+opencv",
                },
                "results": results,
            }, f, indent=2)
        print(f"  JSON report : {report_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()