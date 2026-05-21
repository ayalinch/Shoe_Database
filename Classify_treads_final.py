#!/usr/bin/env python3
"""
Outsole Tread Classifier + Annotator  (v6 - lug fingerprint output)
====================================================================
Adds similarity-style console output per image:
  [view_type]  lugs=N  circ=X.XX  asp=X.XX  mean_px=XX  -> [N%] Pattern

Key fix: extract_signals() now uses CLAHE-enhanced adaptive threshold
for blob/peg/stud detection instead of Otsu, so dark soles (mean<55px)
no longer return lugs=0-6 with near-zero circ/asp.

Panels: same 6-panel dark layout as v5.
"""

import cv2
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR      = Path.home() / "Desktop" / "extracted_boot_images" / "classified"
SOLE_DIR      = BASE_DIR / "sole"
ANNOTATED_DIR = BASE_DIR / "annotated"
EXCEL_PATH    = BASE_DIR / "tread_classification.xlsx"
IMAGE_EXTS    = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}
ANNOTATED_DIR.mkdir(parents=True, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SOLE MASK
# ══════════════════════════════════════════════════════════════════════════════

def find_sole_mask(gray, h, w):
    blur  = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blur, 20, 80)
    k     = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (25, 25))
    closed = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, k)
    filled = cv2.dilate(closed, k, iterations=2)
    cnts, _ = cv2.findContours(filled, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best = None; best_score = 0.0
    for c in cnts:
        area = cv2.contourArea(c)
        if area < h*w*0.08 or area > h*w*0.95: continue
        x, y, cw, ch = cv2.boundingRect(c)
        if ch == 0: continue
        aspect_ok = cw/ch > 1.1
        cx_ = x+cw/2; cy_ = y+ch/2
        centred = abs(cx_/w-0.5)<0.35 and abs(cy_/h-0.5)<0.40
        score = area*(1 if aspect_ok else 0.4)*(1 if centred else 0.5)
        if score > best_score: best_score=score; best=c
    mask = np.zeros((h, w), dtype=np.uint8)
    if best is not None:
        cv2.drawContours(mask, [best], -1, 255, -1)
        mask = cv2.erode(mask, cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(11,11)), iterations=1)
    else:
        # Dark-background fallback (perspective boot photos)
        hsv = cv2.cvtColor(cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR), cv2.COLOR_BGR2HSV)
        fg  = cv2.bitwise_not((hsv[:,:,2] < 35).astype(np.uint8)*255)
        k3  = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(15,15))
        fg  = cv2.morphologyEx(fg, cv2.MORPH_CLOSE, k3)
        n_lab, labels, stats, _ = cv2.connectedComponentsWithStats(fg)
        if n_lab > 1:
            largest = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
            mask = (labels == largest).astype(np.uint8)*255
            mask = cv2.erode(mask, cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(7,7)),1)
            if np.sum(mask>0) > h*w*0.05:
                return mask
        my, mx = int(h*0.05), int(w*0.05)
        mask[my:h-my, mx:w-mx] = 255
    return mask


def _view_type(gray, mask):
    """Guess flat vs perspective from whether dark background exists."""
    bg_frac = np.sum(gray < 35) / gray.size
    return "perspective" if bg_frac > 0.25 else "flat"


# ══════════════════════════════════════════════════════════════════════════════
#  INNER EDGES
# ══════════════════════════════════════════════════════════════════════════════

def get_inner_edges(gray, sole_mask):
    blur  = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blur, 35, 110)
    k     = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (25, 25))
    inner = cv2.erode(sole_mask, k, iterations=1)
    if np.sum(inner > 0) < 200:
        inner = sole_mask
    return cv2.bitwise_and(edges, inner), inner


# ══════════════════════════════════════════════════════════════════════════════
#  BINARY THRESHOLD  (for secondary metrics / void fraction)
# ══════════════════════════════════════════════════════════════════════════════

def threshold_sole(gray, sole_mask):
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    sole_pixels = blur[sole_mask > 0]
    if len(sole_pixels) < 100:
        return np.zeros_like(gray)
    if float(np.mean(sole_pixels)) > 55:
        _, bw = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
        bw = cv2.bitwise_and(bw, sole_mask)
        h, w = gray.shape; cy, cx = h//2, w//2; ph, pw = h//5, w//5
        patch = bw[max(0,cy-ph):cy+ph, max(0,cx-pw):cx+pw]
        pmask = sole_mask[max(0,cy-ph):cy+ph, max(0,cx-pw):cx+pw]
        if np.sum(pmask>0)>0 and np.sum(patch==255)/np.sum(pmask>0)<0.25:
            bw = cv2.bitwise_and(cv2.bitwise_not(
                cv2.threshold(blur,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]), sole_mask)
    else:
        bw = cv2.adaptiveThreshold(blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                    cv2.THRESH_BINARY, 35, 5)
        bw = cv2.bitwise_and(bw, sole_mask)
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    return cv2.morphologyEx(bw, cv2.MORPH_CLOSE, k)


# ══════════════════════════════════════════════════════════════════════════════
#  LUG FINGERPRINT  (same method as sole_similarity.py)
# ══════════════════════════════════════════════════════════════════════════════

def lug_fingerprint(gray, mask):
    """
    CLAHE-enhanced adaptive threshold → contour analysis.
    Returns (n_lugs, mean_circ, mean_asp) matching similarity tool output.
    Works on both light and dark soles.
    """
    clahe  = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8,8))
    enh    = clahe.apply(gray)
    bw     = cv2.adaptiveThreshold(cv2.GaussianBlur(enh,(3,3),0), 255,
                                    cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                    cv2.THRESH_BINARY, 31, 4)
    bw     = cv2.bitwise_and(bw, mask)
    bw     = cv2.morphologyEx(bw, cv2.MORPH_CLOSE,
                               cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3)))
    sole_area = np.sum(mask>0) + 1.0
    cnts, _   = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    circ_v=[]; asp_v=[]
    for c in cnts:
        area = cv2.contourArea(c)
        if area < sole_area*0.0005 or area > sole_area*0.25: continue
        peri = cv2.arcLength(c, True)
        if peri < 1: continue
        circ_v.append(min(4*np.pi*area/peri**2, 1.0))
        x,y,cw,ch = cv2.boundingRect(c)
        asp_v.append(min(cw,ch)/max(cw,ch,1))
    n_lugs   = len(circ_v)
    mean_circ= round(float(np.mean(circ_v)) if circ_v else 0.0, 2)
    mean_asp = round(float(np.mean(asp_v))  if asp_v  else 0.0, 2)
    return n_lugs, mean_circ, mean_asp


# ══════════════════════════════════════════════════════════════════════════════
#  SIGNAL EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_signals(gray, edges_inner, inner_mask):
    h, w = edges_inner.shape
    ia   = np.sum(inner_mask > 0) + 1e-9
    edge_density = np.count_nonzero(edges_inner) / ia
    ec, _ = cv2.findContours(edges_inner, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    n_blobs_raw = len([c for c in ec if cv2.contourArea(c) > 30])

    clahe       = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    enhanced    = clahe.apply(gray)
    blur_e      = cv2.GaussianBlur(enhanced, (5,5), 0)
    edges_clahe = cv2.bitwise_and(cv2.Canny(blur_e, 40, 120), inner_mask)
    sole_mean   = float(np.mean(gray[inner_mask>0])) if ia>100 else 128.0
    use_clahe   = sole_mean < 80

    # Use CLAHE blob count for dark soles — raw Canny misses internal structure
    ec_clahe, _ = cv2.findContours(edges_clahe, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    n_blobs_clahe = len([c for c in ec_clahe if cv2.contourArea(c) > 30])
    n_blobs = n_blobs_clahe if use_clahe else max(n_blobs_raw, n_blobs_clahe)
    edges_for_lines = edges_clahe if use_clahe else edges_inner

    lines = cv2.HoughLinesP(edges_for_lines, 1, np.pi/180, 15,
                             minLineLength=w//16, maxLineGap=5)
    pos45=neg45=hz=vt=0; pos_sh=neg_sh=0
    if lines is not None:
        for x1,y1,x2,y2 in lines[:,0]:
            a = np.degrees(np.arctan2(y2-y1, x2-x1))
            if   25<a<65:           pos45+=1
            elif -65<a<-25:         neg45+=1
            elif abs(a)<18:         hz+=1
            elif abs(abs(a)-90)<18: vt+=1
            if  10<a<30:            pos_sh+=1
            elif -30<a<-10:         neg_sh+=1
    total_lines = pos45+neg45+hz+vt+1e-9
    v_sym    = min(pos45,neg45)/(max(pos45,neg45)+1e-9)
    n_diag   = pos45+neg45
    v_sym_sh = min(pos_sh,neg_sh)/(max(pos_sh,neg_sh)+1e-9)
    n_shallow= pos_sh+neg_sh

    row_sums = (edges_clahe if use_clahe else edges_inner).sum(axis=1)
    period   = 0.0
    if row_sums.sum()>0:
        fft    = np.abs(np.fft.rfft(row_sums))
        period = float(fft[3:15].sum()/(fft[1:].sum()+1e-9))

    local_std = float(np.std(gray[inner_mask>0])) if ia>100 else 0.0

    # ── CLAHE-enhanced adaptive threshold for blob/peg/stud detection ─────────
    # Replaces threshold_sole() which uses Otsu and fails on dark soles.
    clahe_b = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8,8))
    gray_b  = clahe_b.apply(gray)
    bw_b    = cv2.adaptiveThreshold(cv2.GaussianBlur(gray_b,(5,5),0), 255,
                                     cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 31, 4)
    bw_b    = cv2.bitwise_and(bw_b, inner_mask)
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3))
    solid = cv2.morphologyEx(bw_b, cv2.MORPH_OPEN, k)
    pcnts,_ = cv2.findContours(solid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    peg_areas=[]
    for c in pcnts:
        area=cv2.contourArea(c)
        if not(ia*0.0003<area<ia*0.015): continue
        peri=cv2.arcLength(c,True)
        if peri<1: continue
        if 4*np.pi*area/peri**2>0.35: peg_areas.append(area)
    n_pegs=len(peg_areas)
    peg_regularity=(max(0.0,1.0-np.std(peg_areas)/(np.mean(peg_areas)+1e-9)*2)
                    if n_pegs>3 else 0.0)

    studs=[]
    h_gray, w_gray = gray.shape
    sole_cx, sole_cy = w_gray/2, h_gray/2
    # Inner zone where logos live: central 50% width × 40% height
    logo_x0 = w_gray * 0.25; logo_x1 = w_gray * 0.75
    logo_y0 = h_gray * 0.30; logo_y1 = h_gray * 0.70
    for c in pcnts:
        area=cv2.contourArea(c)
        if not(ia*0.0008<area<ia*0.06): continue
        peri=cv2.arcLength(c,True)
        if peri<1: continue
        if 4*np.pi*area/peri**2>0.55:
            (cx,cy),r=cv2.minEnclosingCircle(c)
            # Reject if dead-center (logo zone) — real studs are near perimeter
            if logo_x0 < cx < logo_x1 and logo_y0 < cy < logo_y1:
                continue
            studs.append((int(cx),int(cy),int(r)))
    # Secondary filter: if only 1-3 studs and they're tightly clustered, likely logos
    if 0 < len(studs) <= 4:
        xs = [s[0] for s in studs]; ys = [s[1] for s in studs]
        spread = max(max(xs)-min(xs), max(ys)-min(ys))
        if spread < min(h_gray, w_gray) * 0.25:
            studs = []  # clustered in small area = branding, not traction studs

    return dict(edge_density=edge_density,n_blobs=n_blobs,
                v_sym=v_sym,n_diag=n_diag,pos45=pos45,neg45=neg45,
                hz=hz,vt=vt,total_lines=total_lines,
                period=period,local_std=local_std,
                n_pegs=n_pegs,peg_regularity=peg_regularity,studs=studs,
                v_sym_sh=v_sym_sh,n_shallow=n_shallow,
                pos_sh=pos_sh,neg_sh=neg_sh,
                use_clahe=use_clahe,sole_mean=sole_mean,
                edges_clahe=edges_clahe)


# ══════════════════════════════════════════════════════════════════════════════
#  PATTERN CLASSIFICATION  — 14 categories
# ══════════════════════════════════════════════════════════════════════════════
#
#  Category                     Key discriminators
#  ─────────────────────────────────────────────────────────────────────────
#  Herringbone / Chevron         high vsym + diagonal lines + FFT period
#  Chevron-Lug                   diagonal sym but lug-shaped elements (low circ)
#  Lugged / Block — Large        few big lugs, low circ, high local_std
#  Lugged / Block — Dense        many small lugs, moderate circ
#  Multi-directional Lug         high lug count, mixed circ, omnidirectional
#  Studded / Hardware            high circ studs, detected metal elements
#  Grid / Waffle                 regular peg array, high peg_regularity
#  Linear / Ribbed               dominant hz or vt lines, low void
#  Wave / S-curve                very low ed, smooth curves
#  Hybrid — Zonal                fore ≠ heel pattern
#  Siped / Fine-tread            high ed but tiny elements, high blob count
#  Minimal / Smooth              near-zero ed, featureless
#  Transitional / Mixed          moderate signals in multiple categories
#  ─────────────────────────────────────────────────────────────────────────

def classify_pattern(sig, gray, sole_mask, h, w, lf=None):
    """
    lf : dict with keys n_lugs, mean_circ, mean_asp from lug_fingerprint()
         (passed through from analyze_image). Falls back to zeros if None.
    """
    ed   = sig['edge_density'];  nb   = sig['n_blobs']
    vsym = sig['v_sym'];         nd   = sig['n_diag']
    per  = sig['period'];        np_  = sig['n_pegs']
    preg = sig['peg_regularity'];ns   = len(sig['studs'])
    vsym_sh = sig.get('v_sym_sh', 0.0); n_sh = sig.get('n_shallow', 0)
    local_std = sig.get('local_std', 0.0)
    sole_area = np.sum(sole_mask>0)+1.0

    # Lug fingerprint from CLAHE contour analysis
    n_lugs    = lf['n_lugs']    if lf else 0
    mean_circ = lf['mean_circ'] if lf else 0.0
    mean_asp  = lf['mean_asp']  if lf else 0.0
    # Lug density: lugs per 10k px of sole area
    lug_density = n_lugs / (sole_area / 10000.0 + 1e-9)

    # ── 1. Herringbone / Chevron ──────────────────────────────────────────────
    # Classic tight V-rows — high diagonal symmetry + FFT period
    herr = 0.0
    if nd >= 6 and vsym > 0.40:
        herr = 0.45*vsym + 0.35*min(1.0, per*5.0) + 0.20*min(1.0, ed/0.12)
        if sig['hz'] > nd*1.5: herr *= 0.4          # horizontal lines dominate → not herr
    if n_sh > 25 and vsym_sh > 0.85 and per > 0.18 and ed > 0.065:
        herr_sh = 0.40*vsym_sh + 0.40*min(1.0,per*5.0) + 0.20*min(1.0,n_sh/30.0)
        herr = max(herr, herr_sh)
    # Suppress if lug fingerprint says elements are large/sparse (chevron-lug territory)
    if n_lugs > 0 and n_lugs < 15 and mean_circ < 0.25:
        herr *= 0.6

    # ── 2. Chevron-Lug ────────────────────────────────────────────────────────
    # Elongated lug elements arranged in a chevron/herringbone layout.
    # Shares diagonal symmetry with herr but elements are blocky not fine.
    chev_lug = 0.0
    if nd >= 4 and vsym > 0.35 and ed > 0.04:
        diag_score = 0.40*vsym + 0.30*min(1.0, nd/15.0)
        # Lug fingerprint: elongated (low circ) but present
        lf_score = 0.0
        if n_lugs >= 6:
            lf_score = 0.30*(1-min(mean_circ, 0.6)/0.6)   # reward low circ
        chev_lug = diag_score + lf_score
        chev_lug *= max(0.0, 1.0 - herr*1.2)              # suppress if already herr

    # ── 3. Studded / Hardware ─────────────────────────────────────────────────
    # Metal studs, carbide pins, traction hardware — high circ + detected studs
    stud = 0.0
    if ns >= 4:
        stud = 0.60*min(1.0, ns/12.0) + 0.40*min(1.0, mean_circ*2)
    elif n_lugs >= 5 and mean_circ > 0.55:
        # No studs detected by circle fitting but fingerprint says very round elements
        stud = 0.30*min(1.0, mean_circ*1.5) + 0.20*min(1.0, n_lugs/20.0)

    # ── 4. Grid / Waffle ──────────────────────────────────────────────────────
    # Small uniform pegs in regular grid — high peg_regularity
    grid = 0.55*min(1.0, np_/20.0) + 0.45*preg if np_ >= 8 else 0.0
    # Lug fingerprint boost: many elements, moderate circ, consistent size
    if n_lugs >= 20 and 0.30 < mean_circ < 0.75 and mean_asp > 0.50:
        grid = max(grid, 0.40*min(1.0,n_lugs/40.0) + 0.30*mean_circ + 0.30*preg)

    # ── 5. Linear / Ribbed ────────────────────────────────────────────────────
    # Dominant parallel ribs — hz or vt lines rule
    lin = 0.0
    if sig['total_lines'] > 4 and nb >= 3:
        dom = max(sig['hz'], sig['vt'])
        linearity = dom / sig['total_lines']
        if linearity > 0.60 and ed < 0.12:
            lin = 0.50*linearity + 0.25*(1-ed/0.12) + 0.25*(1-min(nb,12)/12)
    # Lug fingerprint: high asp (elongated in one direction) boosts linear
    if mean_asp > 0.60 and n_lugs >= 5:
        lin = max(lin, 0.30*mean_asp + 0.20*min(1.0,n_lugs/20.0))

    # ── 6. Wave / S-curve ─────────────────────────────────────────────────────
    # Genuinely smooth — very low edge density, no blob structure
    wave = max(0.0, (0.035-ed)/0.035) if ed < 0.035 else 0.0
    if n_lugs > 8: wave *= 0.3   # blobs detected → not really wave

    # ── 7. Minimal / Smooth ───────────────────────────────────────────────────
    smooth = 0.0
    if ed < 0.025 and local_std < 12:
        smooth = 0.70 - ed/0.025*0.30
    elif ed < 0.03:
        smooth = 0.40*(1 - ed/0.03)

    # ── 8. Lugged / Block — Large ─────────────────────────────────────────────
    # Few, large, irregular blocks — low lug count, large per-lug area, low circ
    lug_large = 0.0
    if ed > 0.025 and n_lugs > 0:
        mean_lug_area = sole_area / (n_lugs + 1e-9)
        size_score  = min(1.0, mean_lug_area / (sole_area*0.05))  # saturates at 5% each
        sparse_score= max(0.0, 1.0 - n_lugs/25.0)                 # fewer lugs → higher
        shape_score = max(0.0, 1.0 - mean_circ/0.50)              # blockier → higher
        lug_large = (0.35*min(1.0,ed/0.10)
                   + 0.25*sparse_score
                   + 0.25*shape_score
                   + 0.15*min(1.0,local_std/40.0))
        lug_large *= max(0.0, 1.0 - herr*0.9)
        lug_large *= max(0.0, 1.0 - chev_lug*0.7)

    # ── 9. Lugged / Block — Dense ─────────────────────────────────────────────
    # Many smaller lugs — high lug count, moderate circ
    lug_dense = 0.0
    if ed > 0.03 and n_lugs >= 15:
        density_score = min(1.0, lug_density/8.0)
        lug_dense = (0.40*density_score
                   + 0.30*min(1.0, ed/0.12)
                   + 0.30*min(1.0, mean_circ/0.50))
        lug_dense *= max(0.0, 1.0 - herr*0.8)

    # ── 10. Multi-directional Lug ─────────────────────────────────────────────
    # High lug count, mixed circ (neither very round nor very elongated),
    # low line directionality (omnidirectional)
    multi = 0.0
    if n_lugs >= 20 and ed > 0.04:
        omni_score  = 1.0 - min(vsym, 0.8)/0.8      # low diagonal sym → more omni
        circ_mix    = 1.0 - abs(mean_circ - 0.40)/0.40  # circ near 0.4 = mixed
        multi = (0.35*min(1.0,lug_density/10.0)
               + 0.35*max(0.0,circ_mix)
               + 0.30*max(0.0,omni_score))
        multi *= max(0.0, 1.0 - herr*0.7)
        multi *= max(0.0, 1.0 - grid*0.8)

    # ── 11. Siped / Fine-tread ────────────────────────────────────────────────
    # High edge density but elements are tiny — many cuts/sipes across large lugs
    siped = 0.0
    if ed > 0.10 and nb > 15:
        # Many fine edge blobs, high density
        siped = 0.50*min(1.0,(ed-0.10)/0.10) + 0.50*min(1.0,nb/30.0)
        siped *= max(0.0, 1.0 - herr*0.5)

    # ── 12. Hybrid — Zonal ────────────────────────────────────────────────────
    # Fore and heel have genuinely different patterns
    split = int(h*0.50)
    m_fore = sole_mask[:split,:]; m_heel = sole_mask[split:,:]
    hyb = 0.0; fore_p = "Mixed"; heel_p = "Mixed"
    if np.sum(m_fore>0)>500 and np.sum(m_heel>0)>500:
        def quick_pat(gray_z, mask_z):
            ei, im = get_inner_edges(gray_z, mask_z)
            sg = extract_signals(gray_z, ei, im)
            if sg['n_diag']>=4 and sg['v_sym']>0.40:        return "Herringbone"
            if sg['edge_density']<0.03:                      return "Smooth"
            if sg['edge_density']>0.09:                      return "Dense"
            if max(sg['hz'],sg['vt'])/sg['total_lines']>0.6: return "Linear"
            return "Mixed"
        fore_p = quick_pat(gray[:split,:], m_fore)
        heel_p = quick_pat(gray[split:,:], m_heel)
        if fore_p != heel_p and "Mixed" not in (fore_p, heel_p):
            hyb = 0.70

    # ── 13. Transitional / Mixed ──────────────────────────────────────────────
    # Catch soles that don't cleanly fit anywhere — moderate everything
    transit = 0.0
    all_raw_max = max(herr,chev_lug,stud,grid,lin,wave,smooth,
                      lug_large,lug_dense,multi,siped,hyb)
    if all_raw_max < 0.25 and ed > 0.02:
        transit = max(0.0, 0.25 - all_raw_max)

    # ── Assemble & normalise ──────────────────────────────────────────────────
    raw = {
        "Herringbone / Chevron":       herr,
        "Chevron-Lug":                 chev_lug,
        "Lugged / Block — Large":      lug_large,
        "Lugged / Block — Dense":      lug_dense,
        "Multi-directional Lug":       multi,
        "Studded / Hardware":          stud,
        "Grid / Waffle":               grid,
        "Linear / Ribbed":             lin,
        "Wave / S-curve":              wave,
        "Hybrid — Zonal":              hyb,
        "Siped / Fine-tread":          siped,
        "Minimal / Smooth":            smooth,
        "Transitional / Mixed":        transit,
    }
    total  = sum(raw.values()) + 1e-9
    scores = {k: round(v/total*100) for k,v in raw.items()}
    winner = max(scores, key=scores.get)

    debug = dict(
        ed=round(ed,3), n_blobs=nb, v_sym=round(vsym,2), n_diag=nd,
        period=round(per,3), n_studs=ns, n_pegs=np_,
        lug_large_raw=round(lug_large,3), lug_dense_raw=round(lug_dense,3),
        herr_raw=round(herr,3), chev_raw=round(chev_lug,3),
        multi_raw=round(multi,3), siped_raw=round(siped,3),
        v_sym_sh=round(vsym_sh,2), n_shallow=n_sh,
        n_lugs=n_lugs, mean_circ=round(mean_circ,2), mean_asp=round(mean_asp,2),
        lug_density=round(lug_density,1),
        fore_pat=fore_p, heel_pat=heel_p,
        use_clahe=sig.get('use_clahe',False)
    )
    return winner, scores, sig['studs'], debug


# ══════════════════════════════════════════════════════════════════════════════
#  SECONDARY CLASSIFIERS
# ══════════════════════════════════════════════════════════════════════════════

def classify_tread_element(bw,edges_inner,sole_mask):
    sa=np.sum(sole_mask>0)+1e-9
    vf=np.sum(cv2.bitwise_and(cv2.bitwise_not(bw),sole_mask)>0)/sa
    ed=np.count_nonzero(edges_inner)/sa
    if vf>0.40 and ed>0.07: return "Lug-dominant"
    elif vf<0.20:            return "Groove-dominant"
    else:                    return "Mixed"

def classify_groove_scale(edges_inner,sole_mask):
    d=np.count_nonzero(edges_inner)/(np.sum(sole_mask>0)+1e-9)
    return "Fine" if d>0.18 else "Medium" if d>0.08 else "Coarse"

def classify_void_fraction(bw,sole_mask):
    sa=np.sum(sole_mask>0)+1e-9
    vf=np.sum(cv2.bitwise_and(cv2.bitwise_not(bw),sole_mask)>0)/sa
    pct=int(vf*100)
    if vf<0.25:   return f"Low void (<25%) [{pct}%]"
    elif vf<0.50: return f"Medium void (25-50%) [{pct}%]"
    else:         return f"High void (>50%) [{pct}%]"

def classify_edge_density(edges_inner,sole_mask):
    d=np.count_nonzero(edges_inner)/(np.sum(sole_mask>0)+1e-9)
    return "Low" if d<0.06 else "Medium" if d<0.15 else "High"

def classify_directionality(gray):
    sx=cv2.Sobel(gray,cv2.CV_64F,1,0,ksize=3)
    sy=cv2.Sobel(gray,cv2.CV_64F,0,1,ksize=3)
    a=np.arctan2(np.abs(sy),np.abs(sx))*180/np.pi
    hz=np.sum(a<20)/a.size; vt=np.sum(a>70)/a.size
    dg=np.sum((a>=20)&(a<=70))/a.size; dom=max(hz,vt,dg)
    if dom==dg and dg>0.45: return "Omnidirectional"
    if hz>0.35 and vt>0.25: return "Bidirectional"
    if vt>0.35 and hz>0.25: return "Bidirectional"
    if dom>0.55:             return "Unidirectional"
    return "Omnidirectional"

def classify_motifs(gray,bw,edges_inner,studs,sole_mask,h,w):
    notes=[]; sa=np.sum(sole_mask>0)+1e-9
    n_s=len(studs)
    if n_s>=8:   notes.append(f"Metal nails/studs visible ({n_s}) — likely ice traction device")
    elif n_s>=4: notes.append(f"Circular studs detected ({n_s}) — possible traction hardware")
    elif n_s>=2: notes.append(f"Possible pivot circle or heel studs ({n_s} circular elements)")
    blur=cv2.GaussianBlur(gray,(3,3),0)
    bright_mask=cv2.bitwise_and((blur>200).astype(np.uint8)*255,sole_mask)
    bright_cnts,_=cv2.findContours(bright_mask,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    n_bright=0
    for c in bright_cnts:
        area=cv2.contourArea(c)
        if not(sa*0.0003<area<sa*0.015): continue
        peri=cv2.arcLength(c,True)
        if peri<1: continue
        if 4*np.pi*area/peri**2>0.40: n_bright+=1
    if n_bright>=4 and n_s<2:
        notes.append(f"Bright reflective spots ({n_bright}) — possible metal nail heads")
    mg=int(min(h,w)*0.12); bm=np.zeros_like(bw)
    bm[:mg,:]=255; bm[-mg:,:]=255; bm[:,:mg]=255; bm[:,-mg:]=255
    bc,_=cv2.findContours(cv2.bitwise_and(bw,bm),cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    n_perim=sum(1 for c in bc if cv2.contourArea(c)>h*w*0.003)
    if n_perim>=5:   notes.append("Prominent perimeter lugs around sole edge")
    elif n_perim>=3: notes.append("Perimeter lugs present")
    mid_col=gray[:,int(w*0.35):int(w*0.65)]
    sole_mean=np.mean(gray[sole_mask>0]) if sa>100 else 128
    if np.mean(mid_col)<sole_mean*0.72:
        notes.append("Central longitudinal channel (flex groove)")
    split=int(h*0.5)
    fore_e=edges_inner[:split,:]; heel_e=edges_inner[split:,:]
    fa=np.sum(sole_mask[:split,:]>0)+1e-9; ha=np.sum(sole_mask[split:,:]>0)+1e-9
    fore_ed=np.count_nonzero(fore_e)/fa; heel_ed=np.count_nonzero(heel_e)/ha
    if fore_ed>heel_ed*1.6:   notes.append("Forefoot has denser tread than heel")
    elif heel_ed>fore_ed*1.6: notes.append("Heel has denser tread than forefoot")
    ff_e=edges_inner[:int(h*0.4),:]
    ffl=cv2.HoughLinesP(ff_e,1,np.pi/180,25,minLineLength=w//5,maxLineGap=8)
    if ffl is not None:
        hff=sum(1 for l in ffl[:,0]
                if abs(np.degrees(np.arctan2(l[3]-l[1],l[2]-l[0])))<22)
        if hff>=3: notes.append("Flex grooves across forefoot")
    cnts,_=cv2.findContours(bw,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    sipes=sum(1 for c in cnts
              if cv2.contourArea(c)>h*w*0.01 and
              np.count_nonzero(cv2.bitwise_and(
                  edges_inner,
                  cv2.drawContours(np.zeros_like(gray),[c],-1,255,-1)))
              /(cv2.contourArea(c)+1e-9)>0.15)
    if sipes>=3:   notes.append("Siping visible (fine cuts inside lug blocks)")
    elif sipes>=2: notes.append("Possible siping on some lugs")
    void_frac=np.sum(cv2.bitwise_and(cv2.bitwise_not(bw),sole_mask)>0)/sa
    edge_d=np.count_nonzero(edges_inner)/sa
    if void_frac>0.55: notes.append("Very open tread (>55% void) — deep aggressive lug profile")
    elif void_frac<0.15 and edge_d<0.06: notes.append("Very closed/smooth tread — low grip texture")
    return notes if notes else ["No distinctive features detected"]


# ══════════════════════════════════════════════════════════════════════════════
#  ANNOTATION  (v5 dark panels — unchanged)
# ══════════════════════════════════════════════════════════════════════════════

def _clahe_edges(gray, mask):
    clahe    = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    enhanced = clahe.apply(gray)
    blur     = cv2.GaussianBlur(enhanced, (5,5), 0)
    edges    = cv2.Canny(blur, 35, 110)
    k        = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (21,21))
    inner    = cv2.erode(mask, k, iterations=1)
    if np.sum(inner>0)<300: inner=mask
    return cv2.bitwise_and(edges, inner), inner


def draw_annotations(img, result, studs, sole_mask, edges_inner, inner_mask, bw, scores):
    """
    Clean 2×2 square grid:
      [Original]        [CLAHE edges]
      [Texture heatmap] [Lug circularity]
    Title bar above, slim debug bar below.
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    font = cv2.FONT_HERSHEY_SIMPLEX
    fs   = max(0.28, min(0.40, w/800))

    def label(panel, text, color=(200,220,255)):
        cv2.putText(panel, text, (5,16), font, fs*0.85, (0,0,0), 2, cv2.LINE_AA)
        cv2.putText(panel, text, (5,16), font, fs*0.85, color,   1, cv2.LINE_AA)

    clahe_edges, clahe_inner = _clahe_edges(gray, sole_mask)
    mask_boundary = cv2.Canny(sole_mask, 50, 150)
    sole_area_px  = np.sum(inner_mask>0) + 1.0
    BG = 12

    # ── A: Original ───────────────────────────────────────────────────────────
    pa = img.copy()
    pa[mask_boundary>0] = (100, 50, 15)
    ranked  = sorted(scores.items(), key=lambda x: -x[1])
    win_txt = f"{ranked[0][0]}  {ranked[0][1]}%"
    cv2.putText(pa, result['filename'], (5,16), font, fs*0.80, (0,0,0),      2, cv2.LINE_AA)
    cv2.putText(pa, result['filename'], (5,16), font, fs*0.80, (220,220,220), 1, cv2.LINE_AA)
    cv2.putText(pa, win_txt,            (5,30), font, fs*0.72, (0,0,0),      2, cv2.LINE_AA)
    cv2.putText(pa, win_txt,            (5,30), font, fs*0.72, (0,225,255),  1, cv2.LINE_AA)

    # ── B: CLAHE edges on darkened original ───────────────────────────────────
    pb = (img.astype(np.float32)*0.40).clip(0,255).astype(np.uint8)
    ov = np.zeros_like(pb); ov[clahe_edges>0] = (0,215,255)
    pb = cv2.addWeighted(pb, 1.0, ov, 0.95, 0)
    pb[mask_boundary>0] = (180, 50, 10)
    for (cx_,cy_,r) in studs:
        cv2.circle(pb,(cx_,cy_),r,(0,255,160),2)
        cv2.circle(pb,(cx_,cy_),3,(255,255,255),-1)
    lf_txt = (f"lugs={result.get('_n_lugs',0)}  "
              f"circ={result.get('_mean_circ',0):.2f}  "
              f"asp={result.get('_mean_asp',0):.2f}")
    cv2.putText(pb, lf_txt, (5,h-6), font, fs*0.65, (0,200,255), 1, cv2.LINE_AA)
    label(pb, "Tread edges", (0,215,255))

    # ── C: Texture depth heatmap ──────────────────────────────────────────────
    sole_f  = gray.astype(np.float32); sole_f[sole_mask==0] = 0
    ksize   = max(9, min(21, w//28)) | 1
    mean1   = cv2.blur(sole_f,(ksize,ksize))
    mean2   = cv2.blur(sole_f**2,(ksize,ksize))
    std_map = np.sqrt(np.clip(mean2-mean1**2, 0, None))
    tex     = cv2.normalize(std_map, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    pc      = cv2.applyColorMap(tex, cv2.COLORMAP_INFERNO)
    pc[sole_mask==0]    = (BG,BG,BG)
    pc[mask_boundary>0] = (70,70,70)
    for (cx_,cy_,r) in studs:
        cv2.circle(pc,(cx_,cy_),r,(0,255,200),2)
        cv2.circle(pc,(cx_,cy_),3,(255,255,255),-1)
    label(pc, "Texture depth  (hot=complex)", (255,190,50))

    # ── D: Lug circularity map ────────────────────────────────────────────────
    pd = np.full((h,w,3), BG, dtype=np.uint8)
    pd[mask_boundary>0] = (45,45,55)

    clahe_pd = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8,8))
    blur_pd  = cv2.GaussianBlur(clahe_pd.apply(gray),(3,3),0)
    bw_pd    = cv2.adaptiveThreshold(blur_pd, 255,
                                      cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                      cv2.THRESH_BINARY, 21, 3)
    bw_pd    = cv2.bitwise_and(bw_pd, inner_mask)
    bw_pd    = cv2.morphologyEx(bw_pd, cv2.MORPH_OPEN,
                                 cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3)))
    lug_cnts,_ = cv2.findContours(bw_pd, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)

    drawn = 0
    for c in lug_cnts:
        area = cv2.contourArea(c)
        if area < sole_area_px*0.0002 or area > sole_area_px*0.40: continue
        peri = cv2.arcLength(c,True)
        if peri < 1: continue
        cv_ = min(4*np.pi*area/peri**2, 1.0)
        cv2.drawContours(pd,[c],-1,(0,int(cv_*255),int((1-cv_)*255)),cv2.FILLED)
        cv2.drawContours(pd,[c],-1,(50,50,50),1)
        drawn += 1

    pd[mask_boundary>0] = (45,45,55)
    # compact legend bottom-right
    lx = w-118
    for yi,txt,col in [(h-26,"● round",(0,220,80)),
                        (h-14,"● elongated",(0,60,220))]:
        cv2.putText(pd,txt,(lx,yi),font,0.28,col,1,cv2.LINE_AA)
    cv2.putText(pd,f"{drawn} elements",(5,h-6),font,0.28,(120,120,120),1,cv2.LINE_AA)
    label(pd,"Lug circularity",(180,210,255))

    # ── Assemble 2×2 ─────────────────────────────────────────────────────────
    row1  = np.hstack([pa, pb])
    row2  = np.hstack([pc, pd])
    grid  = np.vstack([row1, row2])
    W     = grid.shape[1]

    # ── Title bar ─────────────────────────────────────────────────────────────
    top2 = f"  2nd: {ranked[1][0]} ({ranked[1][1]}%)" if ranked[1][1]>6 else ""
    meta = (f"  |  {result['B_element']}"
            f"  |  {result['D_void'].split('[')[0].strip()}"
            f"  |  {result['F_directionality']}"
            f"  |  {result['filename']}")
    title_txt = f"[{ranked[0][1]}%] {ranked[0][0]}" + top2 + meta
    tbar = np.full((26,W,3), 10, dtype=np.uint8)
    cv2.putText(tbar,title_txt,(6,19),font,0.42,(0,0,0),     2,cv2.LINE_AA)
    cv2.putText(tbar,title_txt,(6,19),font,0.42,(0,225,255), 1,cv2.LINE_AA)

    # ── Debug bar ─────────────────────────────────────────────────────────────
    dbg = result['_debug']
    dbg_txt = (f"ed={dbg.get('ed',0):.3f}  blobs={dbg.get('n_blobs',0)}"
               f"  vsym={dbg.get('v_sym',0):.2f}  nd={dbg.get('n_diag',0)}"
               f"  lugs={dbg.get('n_lugs',0)}  dens={dbg.get('lug_density',0):.1f}"
               f"  circ={dbg.get('mean_circ',0):.2f}  asp={dbg.get('mean_asp',0):.2f}"
               f"  herr={dbg.get('herr_raw',0):.2f}  chev={dbg.get('chev_raw',0):.2f}"
               f"  Ll={dbg.get('lug_large_raw',0):.2f}  Ld={dbg.get('lug_dense_raw',0):.2f}"
               f"  multi={dbg.get('multi_raw',0):.2f}"
               f"{'  [CLAHE]' if dbg.get('use_clahe') else ''}")
    dbar = np.full((18,W,3), 8, dtype=np.uint8)
    cv2.putText(dbar,dbg_txt,(6,13),font,0.30,(130,130,130),1,cv2.LINE_AA)

    return np.vstack([tbar, grid, dbar])
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    font = cv2.FONT_HERSHEY_SIMPLEX
    fs   = max(0.30, min(0.42, w/800))

    def label(panel, text, color=(200,220,255), y=18):
        cv2.putText(panel, text, (6, y), font, fs*0.85, (0,0,0), 2, cv2.LINE_AA)
        cv2.putText(panel, text, (6, y), font, fs*0.85, color,   1, cv2.LINE_AA)

    clahe_edges, clahe_inner = _clahe_edges(gray, sole_mask)
    sole_mean     = float(np.mean(gray[inner_mask>0])) if np.sum(inner_mask>0)>100 else 128.0
    mask_boundary = cv2.Canny(sole_mask, 50, 150)
    sole_area_px  = np.sum(inner_mask>0) + 1.0
    BG = 14  # dark background value

    # ── Panel A: Original image (clean, no overlays) ──────────────────────────
    pa = img.copy()
    # Subtle sole boundary in dim blue so you can see the mask
    pa[mask_boundary>0] = (120, 60, 20)
    # Filename + key stats top-left
    ranked = sorted(scores.items(), key=lambda x: -x[1])
    winner_txt = f"{ranked[0][0]}  {ranked[0][1]}%"
    cv2.putText(pa, result['filename'], (6,18), font, fs*0.80, (0,0,0),     2, cv2.LINE_AA)
    cv2.putText(pa, result['filename'], (6,18), font, fs*0.80, (230,230,230),1, cv2.LINE_AA)
    cv2.putText(pa, winner_txt,         (6,34), font, fs*0.72, (0,0,0),     2, cv2.LINE_AA)
    cv2.putText(pa, winner_txt,         (6,34), font, fs*0.72, (0,230,255), 1, cv2.LINE_AA)

    # ── Panel B: CLAHE edges on darkened original ─────────────────────────────
    pb = (img.astype(np.float32) * 0.45).clip(0,255).astype(np.uint8)
    ov = np.zeros_like(pb)
    ov[clahe_edges>0] = (0, 215, 255)          # yellow-cyan edges
    pb = cv2.addWeighted(pb, 1.0, ov, 0.95, 0)
    pb[mask_boundary>0] = (200, 60, 10)        # dim blue-red boundary
    # Stud circles
    for (cx_, cy_, r) in studs:
        cv2.circle(pb, (cx_,cy_), r,   (0,255,160), 2)
        cv2.circle(pb, (cx_,cy_), 3,   (255,255,255), -1)
    n_lugs = result.get('_n_lugs', 0)
    circ   = result.get('_mean_circ', 0.0)
    asp    = result.get('_mean_asp', 0.0)
    cv2.putText(pb, f"lugs={n_lugs}  circ={circ:.2f}  asp={asp:.2f}",
                (6, h-8), font, fs*0.68, (0,200,255), 1, cv2.LINE_AA)
    label(pb, "Tread edges (CLAHE)", (0,215,255))

    # ── Panel C: Texture depth heatmap (INFERNO) ──────────────────────────────
    sole_f  = gray.astype(np.float32); sole_f[sole_mask==0] = 0
    ksize   = max(9, min(21, w//28)) | 1
    mean1   = cv2.blur(sole_f, (ksize,ksize))
    mean2   = cv2.blur(sole_f**2, (ksize,ksize))
    std_map = np.sqrt(np.clip(mean2 - mean1**2, 0, None))
    tex     = cv2.normalize(std_map, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    pc      = cv2.applyColorMap(tex, cv2.COLORMAP_INFERNO)
    pc[sole_mask==0]    = (BG, BG, BG)
    pc[mask_boundary>0] = (80, 80, 80)
    for (cx_,cy_,r) in studs:
        cv2.circle(pc, (cx_,cy_), r, (0,255,200), 2)
        cv2.circle(pc, (cx_,cy_), 3, (255,255,255), -1)
    label(pc, "Texture depth  (hot=complex)", (255,190,50))

    # ── Panel D: Lug circularity map ──────────────────────────────────────────
    pd = np.full((h,w,3), BG, dtype=np.uint8)
    pd[mask_boundary>0] = (50, 50, 60)

    clahe_pd = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8,8))
    blur_pd  = cv2.GaussianBlur(clahe_pd.apply(gray), (3,3), 0)
    bw_pd    = cv2.adaptiveThreshold(blur_pd, 255,
                                      cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                      cv2.THRESH_BINARY, 21, 3)
    bw_pd    = cv2.bitwise_and(bw_pd, inner_mask)
    bw_pd    = cv2.morphologyEx(bw_pd, cv2.MORPH_OPEN,
                                 cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3)))
    lug_cnts, _ = cv2.findContours(bw_pd, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)

    drawn = 0
    for c in lug_cnts:
        area = cv2.contourArea(c)
        if area < sole_area_px*0.0002 or area > sole_area_px*0.40: continue
        peri = cv2.arcLength(c, True)
        if peri < 1: continue
        circ_v = min(4*np.pi*area/peri**2, 1.0)
        g  = int(circ_v * 255)
        r_ = int((1-circ_v) * 255)
        cv2.drawContours(pd, [c], -1, (0, g, r_), cv2.FILLED)
        cv2.drawContours(pd, [c], -1, (60, 60, 60), 1)
        drawn += 1

    pd[mask_boundary>0] = (50, 50, 60)
    # Compact legend bottom-right
    leg_x = w - 130
    for yi, txt, col in [(h-38, "round (green)", (0,220,80)),
                          (h-24, "mixed (yellow)",(160,220,80)),
                          (h-10, "elongated (red)",(0,60,220))]:
        cv2.putText(pd, txt, (leg_x,yi), font, 0.28, col, 1, cv2.LINE_AA)
    cv2.putText(pd, f"{drawn} elements", (6,h-8), font, 0.30, (140,140,140), 1, cv2.LINE_AA)
    label(pd, "Lug circularity", (180,210,255))

    # ── Panel E: Score bars + metrics ─────────────────────────────────────────
    pe = np.full((h,w,3), BG, dtype=np.uint8)

    bar_colors = {
        "Herringbone / Chevron":   (0,220,255),
        "Chevron-Lug":             (0,180,220),
        "Lugged / Block — Large":  (0,200,100),
        "Lugged / Block — Dense":  (0,160,80),
        "Multi-directional Lug":   (0,230,150),
        "Studded / Hardware":      (60,120,255),
        "Grid / Waffle":           (255,180,0),
        "Linear / Ribbed":         (120,200,255),
        "Wave / S-curve":          (200,80,255),
        "Hybrid — Zonal":          (255,120,50),
        "Siped / Fine-tread":      (180,230,255),
        "Minimal / Smooth":        (150,150,150),
        "Transitional / Mixed":    (100,100,100),
    }

    ranked_all = sorted(scores.items(), key=lambda x: -x[1])
    bar_x = 6; bar_y = 22; bar_h = 14; bar_max_w = w - 14
    for name, pct in ranked_all:
        if pct == 0: continue
        col = bar_colors.get(name, (180,180,180))
        filled_w = int(bar_max_w * pct / 100)
        # Bar background (dim)
        cv2.rectangle(pe, (bar_x,bar_y), (bar_x+bar_max_w, bar_y+bar_h), (30,30,30), -1)
        # Filled portion
        cv2.rectangle(pe, (bar_x,bar_y), (bar_x+filled_w, bar_y+bar_h), col, -1)
        # Label
        txt = f"{name}  {pct}%"
        cv2.putText(pe, txt, (bar_x+4, bar_y+10), font, 0.31, (0,0,0),     2, cv2.LINE_AA)
        cv2.putText(pe, txt, (bar_x+4, bar_y+10), font, 0.31, (255,255,255),1, cv2.LINE_AA)
        bar_y += bar_h + 3

    # Metric readout
    dbg = result['_debug']
    bar_y += 6
    metrics = [
        ("ed",        f"Edge density  {dbg.get('ed',0):.3f}"),
        ("blobs",     f"Blobs         {dbg.get('n_blobs',0)}"),
        ("vsym",      f"V-sym / diag  {dbg.get('v_sym',0):.2f} / {dbg.get('n_diag',0)}"),
        ("lugs",      f"Lugs (lf)     {dbg.get('n_lugs',0)}  dens={dbg.get('lug_density',0):.1f}"),
        ("circ",      f"Circ / Asp    {dbg.get('mean_circ',0):.2f} / {dbg.get('mean_asp',0):.2f}"),
        ("period",    f"FFT period    {dbg.get('period',0):.3f}"),
        ("void",      f"Void frac     {result['D_void'].split('[')[0].strip()}"),
        ("dir",       f"Direction     {result['F_directionality']}"),
    ]
    for _, txt in metrics:
        if bar_y > h - 12: break
        cv2.putText(pe, txt, (bar_x, bar_y), font, 0.30, (140,210,140), 1, cv2.LINE_AA)
        bar_y += 13

    label(pe, "Pattern scores", (200,200,255))

    # ── Assemble single row ───────────────────────────────────────────────────
    strip = np.hstack([pa, pb, pc, pd, pe])
    W = strip.shape[1]

    # ── Title bar ─────────────────────────────────────────────────────────────
    ranked = sorted(scores.items(), key=lambda x: -x[1])
    top1   = f"[{ranked[0][1]}%] {ranked[0][0]}"
    top2   = f"  2nd: {ranked[1][0]} ({ranked[1][1]}%)" if ranked[1][1] > 6 else ""
    meta   = (f"  |  {result['B_element']}"
              f"  |  {result['D_void'].split('[')[0].strip()}"
              f"  |  {result['F_directionality']}"
              f"  |  {result['filename']}")
    title_txt = top1 + top2 + meta

    title_bar = np.full((28, W, 3), 10, dtype=np.uint8)
    cv2.putText(title_bar, title_txt, (8,20), font, 0.44, (0,0,0),     2, cv2.LINE_AA)
    cv2.putText(title_bar, title_txt, (8,20), font, 0.44, (0,225,255), 1, cv2.LINE_AA)

    # ── Debug bar ─────────────────────────────────────────────────────────────
    dbg_txt = (f"ed={dbg.get('ed',0):.3f}  blobs={dbg.get('n_blobs',0)}"
               f"  vsym={dbg.get('v_sym',0):.2f}  nd={dbg.get('n_diag',0)}"
               f"  lugs={dbg.get('n_lugs',0)}  dens={dbg.get('lug_density',0):.1f}"
               f"  circ={dbg.get('mean_circ',0):.2f}  asp={dbg.get('mean_asp',0):.2f}"
               f"  herr={dbg.get('herr_raw',0):.2f}  chev={dbg.get('chev_raw',0):.2f}"
               f"  Ll={dbg.get('lug_large_raw',0):.2f}  Ld={dbg.get('lug_dense_raw',0):.2f}"
               f"  multi={dbg.get('multi_raw',0):.2f}"
               f"{'  [CLAHE]' if dbg.get('use_clahe') else ''}")
    debug_bar = np.full((20, W, 3), 8, dtype=np.uint8)
    cv2.putText(debug_bar, dbg_txt, (8,14), font, 0.33, (140,140,140), 1, cv2.LINE_AA)

    return np.vstack([title_bar, strip, debug_bar])


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

RUBRIC_COLS=["Filename","View","A) Primary Pattern","Conf%","B) Tread Element",
             "C) Groove Scale","D) Void Fraction","E) Edge Density","F) Directionality",
             "Lugs (n)","Circ","Asp","Sole px","G) Key Motifs"]
PATTERN_COLORS={
    "Herringbone / Chevron":     "FFF2CC",
    "Chevron-Lug":               "FFE0A0",
    "Lugged / Block — Large":    "E2EFDA",
    "Lugged / Block — Dense":    "C6EFCE",
    "Multi-directional Lug":     "A9D18E",
    "Studded / Hardware":        "F4B8B8",
    "Grid / Waffle":             "FCE4D6",
    "Linear / Ribbed":           "D9D2E9",
    "Wave / S-curve":            "EAD1DC",
    "Hybrid — Zonal":            "CFE2F3",
    "Siped / Fine-tread":        "DAE3F3",
    "Minimal / Smooth":          "F4CCCC",
    "Transitional / Mixed":      "EDEDED",
}

def build_excel(all_results):
    wb=Workbook(); ws=wb.active; ws.title="Tread Classification"
    HF=PatternFill("solid",start_color="1F3864")
    HFnt=Font(bold=True,color="FFFFFF",name="Arial",size=10)
    NFnt=Font(name="Arial",size=9)
    C=Alignment(horizontal="center",vertical="center",wrap_text=True)
    L=Alignment(horizontal="left",vertical="center",wrap_text=True)
    for col,h_ in enumerate(RUBRIC_COLS,1):
        cell=ws.cell(row=1,column=col,value=h_); cell.font=HFnt; cell.fill=HF; cell.alignment=C
    ws.row_dimensions[1].height=30
    for i,r in enumerate(all_results,2):
        color=next((v for k,v in PATTERN_COLORS.items() if k.lower() in r["A_pattern"].lower()),"FFFFFF")
        rf=PatternFill("solid",start_color=color)
        conf=r.get("_scores",{}).get(r["A_pattern"],"?")
        vals=[r["filename"], r.get("_view_type","?"), r["A_pattern"], conf,
              r["B_element"], r["C_scale"], r["D_void"], r["E_edge_density"],
              r["F_directionality"],
              r.get("_n_lugs","?"), r.get("_mean_circ","?"), r.get("_mean_asp","?"),
              r.get("_sole_mean","?"),
              ", ".join(r["G_motifs"])]
        for col,val in enumerate(vals,1):
            cell=ws.cell(row=i,column=col,value=val); cell.font=NFnt; cell.fill=rf
            cell.alignment=L if col in(1,14) else C
        ws.row_dimensions[i].height=18
    for col,wid in enumerate([32,10,24,7,18,12,20,12,16,8,7,7,8,44],1):
        ws.column_dimensions[get_column_letter(col)].width=wid
    ws.freeze_panes="A2"
    ls=wb.create_sheet("Legend")
    ls.column_dimensions["A"].width=28; ls.column_dimensions["B"].width=52
    b9=Font(bold=True,name="Arial",size=9); n9=Font(name="Arial",size=9)
    ls["A1"]="Tread Classification Legend"
    ls["A1"].font=Font(bold=True,name="Arial",size=11,color="1F3864")
    for row,lbl_,desc,color in [
        (3, "Herringbone / Chevron",   "Tight repeating V-rows; symmetric +45°/−45° grooves; fine tread elements",      "FFF2CC"),
        (4, "Chevron-Lug",             "V-shaped or angled lug blocks; diagonal symmetry but coarser than herringbone",  "FFE0A0"),
        (5, "Lugged / Block — Large",  "Few large separated raised blocks; deep wide channels; low lug count (<20)",     "E2EFDA"),
        (6, "Lugged / Block — Dense",  "Many smaller lug blocks packed across sole; moderate circularity",               "C6EFCE"),
        (7, "Multi-directional Lug",   "High lug count, mixed shapes pointing in multiple directions; omnidirectional",  "A9D18E"),
        (8, "Studded / Hardware",      "Circular studs, carbide pins, or metal traction hardware; very high circularity","F4B8B8"),
        (9, "Grid / Waffle",           "Small uniform round/square pegs in regular repeating grid pattern",              "FCE4D6"),
        (10,"Linear / Ribbed",         "Parallel ribs running heel-to-toe or across width; strong unidirectional lines", "D9D2E9"),
        (11,"Wave / S-curve",          "Smooth flowing curved grooves; very low edge density; minimal void",             "EAD1DC"),
        (12,"Hybrid — Zonal",          "Two distinct pattern zones (forefoot ≠ heel) each covering >35% of sole",        "CFE2F3"),
        (13,"Siped / Fine-tread",      "Dense fine sipes or micro-cuts across lug surface; very high edge density",      "DAE3F3"),
        (14,"Minimal / Smooth",        "Near-featureless sole; very shallow or no visible groove structure",              "F4CCCC"),
        (15,"Transitional / Mixed",    "Moderate signals across multiple categories; does not cleanly fit one type",      "EDEDED"),
    ]:
        a=ls.cell(row=row,column=1,value=lbl_); a.font=b9
        b=ls.cell(row=row,column=2,value=desc); b.font=n9
        f=PatternFill("solid",start_color=color); a.fill=f; b.fill=f
    wb.save(str(EXCEL_PATH)); print(f"  Excel saved: {EXCEL_PATH}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def analyze_image(img_path):
    img=cv2.imread(str(img_path))
    if img is None: return None,None
    h0,w0=img.shape[:2]
    if w0>800: img=cv2.resize(img,(800,int(h0*800/w0)))
    gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    h,w=gray.shape

    sole_mask             =find_sole_mask(gray,h,w)
    edges_inner,inner_mask=get_inner_edges(gray,sole_mask)
    sig                   =extract_signals(gray,edges_inner,inner_mask)
    bw                    =threshold_sole(gray,sole_mask)

    # Lug fingerprint must come BEFORE classify_pattern (passes lf= dict in)
    n_lugs,mean_circ,mean_asp = lug_fingerprint(gray, inner_mask)
    sole_mean_px = round(float(np.mean(gray[inner_mask>0]))
                         if np.sum(inner_mask>0)>100 else 0.0, 1)
    vtype = _view_type(gray, sole_mask)

    pattern,scores,studs,debug=classify_pattern(sig,gray,sole_mask,h,w,
                                                  lf={'n_lugs':n_lugs,'mean_circ':mean_circ,'mean_asp':mean_asp})

    result={
        "filename":         img_path.name,
        "A_pattern":        pattern,
        "B_element":        classify_tread_element(bw,edges_inner,sole_mask),
        "C_scale":          classify_groove_scale(edges_inner,sole_mask),
        "D_void":           classify_void_fraction(bw,sole_mask),
        "E_edge_density":   classify_edge_density(edges_inner,sole_mask),
        "F_directionality": classify_directionality(gray),
        "G_motifs":         classify_motifs(gray,bw,edges_inner,studs,sole_mask,h,w),
        "_scores":          scores,
        "_debug":           debug,
        "_n_lugs":          n_lugs,
        "_mean_circ":       mean_circ,
        "_mean_asp":        mean_asp,
        "_sole_mean":       sole_mean_px,
        "_view_type":       vtype,
    }
    annotated=draw_annotations(img,result,studs,sole_mask,edges_inner,inner_mask,bw,scores)
    return result,annotated


def main():
    if not SOLE_DIR.exists():
        print(f"ERROR: {SOLE_DIR} not found."); return
    images=[p for p in sorted(SOLE_DIR.iterdir())
            if p.suffix.lower() in IMAGE_EXTS and p.is_file()]
    if not images: print(f"No images in {SOLE_DIR}"); return
    print(f"\n{'='*72}")
    print(f"  Tread Classifier v6")
    print(f"  Input : {SOLE_DIR}   ({len(images)} images)")
    print(f"{'='*72}\n")
    all_results=[]
    for img_path in images:
        print(f"  {img_path.name:<42}", end=" ", flush=True)
        result,annotated=analyze_image(img_path)
        if result is None: print("SKIP"); continue
        cv2.imwrite(str(ANNOTATED_DIR/img_path.name),annotated)
        conf  = result["_scores"].get(result["A_pattern"],"?")
        vtype = result["_view_type"]
        nlug  = result["_n_lugs"]
        circ  = result["_mean_circ"]
        asp   = result["_mean_asp"]
        sol   = result["_sole_mean"]
        print(f"[{vtype:11s}]  lugs={nlug:3d}  circ={circ:.2f}  asp={asp:.2f}  "
              f"px={sol:5.1f}  -> [{conf}%] {result['A_pattern']}")
        all_results.append(result)
    if all_results: build_excel(all_results)
    print(f"\n  Done! {len(all_results)} soles classified.\n")


if __name__=="__main__":
    main()